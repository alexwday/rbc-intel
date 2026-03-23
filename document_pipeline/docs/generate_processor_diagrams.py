"""Generate processor flow diagrams as a styled XLSX workbook.

Creates one sheet per processor (PDF, DOCX, PPTX, XLSX) plus a
shared Vision Pipeline sheet. Each sheet shows the processing
flow from raw input to standardized output, marking LLM vs
programmatic decision points.

Usage:
    cd document_pipeline
    .venv/bin/python docs/generate_processor_diagrams.py
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

OUTPUT = Path(__file__).parent / "processor_flow_diagrams.xlsx"

# ── Colors ─────────────────────────────────────────────────────

FILLS = {
    "title": PatternFill(start_color="1E3A5F", fill_type="solid"),
    "input": PatternFill(start_color="DBEAFE", fill_type="solid"),
    "programmatic": PatternFill(start_color="E2E8F0", fill_type="solid"),
    "llm": PatternFill(start_color="FEF3C7", fill_type="solid"),
    "decision": PatternFill(start_color="FDE68A", fill_type="solid"),
    "output": PatternFill(start_color="D1FAE5", fill_type="solid"),
    "error": PatternFill(start_color="FEE2E2", fill_type="solid"),
    "arrow": PatternFill(start_color="F8FAFC", fill_type="solid"),
    "legend_header": PatternFill(start_color="334155", fill_type="solid"),
}
FONTS = {
    "title": Font(name="Calibri", size=14, bold=True, color="FFFFFF"),
    "header": Font(name="Calibri", size=11, bold=True),
    "body": Font(name="Calibri", size=10),
    "small": Font(name="Calibri", size=9, color="64748B"),
    "arrow": Font(name="Calibri", size=12, bold=True, color="475569"),
    "legend": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
}
THIN = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _cell(ws, row, col, value, fill=None, font=None, alignment=None):
    """Write and style a single cell."""
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    cell.alignment = alignment or LEFT
    cell.border = THIN
    return cell


def _merged_cell(ws, row, col, end_col, value, fill=None, font=None):
    """Write a merged cell spanning columns."""
    ws.merge_cells(
        start_row=row, start_column=col,
        end_row=row, end_column=end_col,
    )
    cell = _cell(ws, row, col, value, fill, font, CENTER)
    for c in range(col + 1, end_col + 1):
        ws.cell(row=row, column=c).border = THIN
        if fill:
            ws.cell(row=row, column=c).fill = fill
    return cell


def _arrow_row(ws, row, col, end_col):
    """Draw a downward arrow row."""
    _merged_cell(ws, row, col, end_col, "\u2193", FILLS["arrow"], FONTS["arrow"])


def _step(ws, row, col, end_col, label, detail, step_type="programmatic"):
    """Draw a processing step with label and detail."""
    fill = FILLS.get(step_type, FILLS["programmatic"])
    font = FONTS["header"]
    _merged_cell(ws, row, col, end_col, label, fill, font)
    if detail:
        _merged_cell(
            ws, row + 1, col, end_col, detail,
            fill, FONTS["small"],
        )
        return row + 2
    return row + 1


def _decision(ws, row, col, end_col, question, yes_label, no_label):
    """Draw a decision diamond (as a row)."""
    _merged_cell(
        ws, row, col, end_col,
        f"\u25c7 {question}",
        FILLS["decision"], FONTS["header"],
    )
    ws.cell(row=row + 1, column=col).value = f"  YES \u2192 {yes_label}"
    ws.cell(row=row + 1, column=col).font = FONTS["small"]
    ws.cell(row=row + 1, column=col).fill = FILLS["decision"]
    ws.cell(row=row + 1, column=col).border = THIN
    mid = (col + end_col) // 2 + 1
    ws.merge_cells(
        start_row=row + 1, start_column=col,
        end_row=row + 1, end_column=mid - 1,
    )
    ws.cell(row=row + 1, column=mid).value = f"  NO \u2192 {no_label}"
    ws.cell(row=row + 1, column=mid).font = FONTS["small"]
    ws.cell(row=row + 1, column=mid).fill = FILLS["decision"]
    ws.cell(row=row + 1, column=mid).border = THIN
    ws.merge_cells(
        start_row=row + 1, start_column=mid,
        end_row=row + 1, end_column=end_col,
    )
    for c in range(col, end_col + 1):
        ws.cell(row=row + 1, column=c).border = THIN
    return row + 2


def _legend(ws, row, col):
    """Draw the color legend."""
    _cell(ws, row, col, "LEGEND", FILLS["legend_header"], FONTS["legend"], CENTER)
    _cell(ws, row, col + 1, "", FILLS["legend_header"])
    items = [
        ("input", "Input / Raw File"),
        ("programmatic", "Programmatic Step"),
        ("llm", "LLM Decision / Call"),
        ("decision", "Decision Point"),
        ("output", "Output"),
        ("error", "Error / Failure"),
    ]
    for i, (key, label) in enumerate(items):
        _cell(ws, row + 1 + i, col, "", FILLS[key])
        _cell(ws, row + 1 + i, col + 1, label, FILLS[key], FONTS["body"])


def _set_col_widths(ws, widths):
    """Set column widths."""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── PDF Sheet ──────────────────────────────────────────────────


def _build_pdf_sheet(ws):
    ws.title = "PDF Processor"
    _set_col_widths(ws, [4, 30, 30, 30, 4])
    c1, c2 = 2, 4
    r = 1

    _merged_cell(ws, r, c1, c2, "PDF PROCESSOR FLOW", FILLS["title"], FONTS["title"])
    r += 1
    _merged_cell(ws, r, c1, c2,
                 "Input: .pdf file  |  Output: ExtractionResult with PageResult per page",
                 FILLS["title"], FONTS["small"])
    r += 2

    r = _step(ws, r, c1, c2, "1. INPUT: Raw PDF File",
              "file_path passed to process_pdf()", "input")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "2. Open PDF [PROGRAMMATIC]",
              "PyMuPDF (fitz) opens document, builds render matrix at configured DPI scale")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "3. FOR EACH PAGE (sequential):",
              "Pages processed 1 to N in order; previous page content carried forward", "programmatic")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "4. Render Page to PNG [PROGRAMMATIC]",
              "fitz renders page at DPI scale (default 2x = 144 DPI) \u2192 PNG bytes")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _decision(ws, r, c1, c2, "Is this page 2+? (previous content exists)",
                  "Augment prompt with 800-char tail of previous page",
                  "Use base extraction prompt")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "5. Vision Extraction [LLM \u2b50]",
              "process_page() sends PNG to LLM with 4-tier fallback chain "
              "(see Vision Pipeline sheet). Returns page_title + markdown content.", "llm")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _decision(ws, r, c1, c2, "Is this page 2+?",
                  "Call page continuation classifier",
                  "Use default metadata (all false)")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "6. Page Continuation Classification [LLM \u2b50]",
              "Sends head+tail of current page + tail of previous page to LLM. "
              "Returns: continued_from_previous_page, table_continuation_detected, "
              "repeated_header_detected, repeated_footer_detected", "llm")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "7. Derive Secondary Flags [PROGRAMMATIC]",
              "contains_page_furniture = repeated_header OR repeated_footer "
              "(simple boolean combination of LLM outputs)")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "8. Build PageResult [PROGRAMMATIC]",
              "PageResult(page_number, page_title, content, method, metadata)")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "9. OUTPUT: ExtractionResult",
              "file_path, filetype='pdf', pages=[PageResult, ...], "
              "total_pages, pages_succeeded, pages_failed=0", "output")
    r += 1

    _merged_cell(ws, r, c1, c2, "ERROR HANDLING: All-or-nothing. "
                 "Any page failure aborts the entire file (RuntimeError raised).",
                 FILLS["error"], FONTS["small"])
    r += 2

    _legend(ws, r, c1)


# ── DOCX Sheet ─────────────────────────────────────────────────


def _build_docx_sheet(wb):
    ws = wb.create_sheet("DOCX Processor")
    _set_col_widths(ws, [4, 30, 30, 30, 4])
    c1, c2 = 2, 4
    r = 1

    _merged_cell(ws, r, c1, c2, "DOCX PROCESSOR FLOW", FILLS["title"], FONTS["title"])
    r += 1
    _merged_cell(ws, r, c1, c2,
                 "Input: .docx file  |  Output: ExtractionResult with PageResult per page",
                 FILLS["title"], FONTS["small"])
    r += 2

    r = _step(ws, r, c1, c2, "1. INPUT: Raw DOCX File",
              "file_path passed to process_docx()", "input")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "2. Convert DOCX \u2192 PDF [PROGRAMMATIC]",
              "LibreOffice headless subprocess converts to PDF. "
              "Isolated user profile, 120s timeout, thread-locked. "
              "Requires metrically-compatible fonts for accurate page breaks.")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "3. Open PDF + Render Pages [PROGRAMMATIC]",
              "Same as PDF processor: PyMuPDF opens converted PDF, renders each page to PNG")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "4. FOR EACH PAGE: Vision Extraction [LLM \u2b50]",
              "Same 4-tier fallback as PDF. Context from previous page passed for pages 2+.", "llm")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "5. Page Continuation Classification [LLM \u2b50]",
              "Same as PDF processor. Additional flag: section_continuation_detected = "
              "continued AND NOT table_continuation (derived programmatically from LLM output).", "llm")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "6. OUTPUT: ExtractionResult",
              "filetype='docx', pages=[PageResult, ...], pages_failed=0", "output")
    r += 2

    _merged_cell(ws, r, c1, c2,
                 "KEY DIFFERENCE FROM PDF: The LibreOffice conversion step. "
                 "Page count may differ from the original Word rendering due to "
                 "font substitution affecting line wrapping and page breaks.",
                 FILLS["programmatic"], FONTS["small"])
    r += 2
    _legend(ws, r, c1)


# ── PPTX Sheet ─────────────────────────────────────────────────


def _build_pptx_sheet(wb):
    ws = wb.create_sheet("PPTX Processor")
    _set_col_widths(ws, [4, 30, 30, 30, 4])
    c1, c2 = 2, 4
    r = 1

    _merged_cell(ws, r, c1, c2, "PPTX PROCESSOR FLOW", FILLS["title"], FONTS["title"])
    r += 1
    _merged_cell(ws, r, c1, c2,
                 "Input: .pptx file  |  Output: ExtractionResult with PageResult per slide",
                 FILLS["title"], FONTS["small"])
    r += 2

    r = _step(ws, r, c1, c2, "1. INPUT: Raw PPTX File",
              "file_path passed to process_pptx()", "input")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "2. Convert PPTX \u2192 PDF [PROGRAMMATIC]",
              "LibreOffice headless conversion (same as DOCX). "
              "Each slide becomes one PDF page.")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "3. Open PDF + Render Slides [PROGRAMMATIC]",
              "PyMuPDF renders each slide to PNG at configured DPI")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "4. FOR EACH SLIDE (sequential, no context passing):",
              "Unlike PDF/DOCX, slides are independent \u2014 no previous-page context. "
              "Each slide gets the same base extraction prompt.", "programmatic")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "5. Vision Extraction [LLM \u2b50]",
              "process_page() with 4-tier fallback. "
              "Extracts slide title + structured markdown content.", "llm")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "6. Slide Classification [LLM \u2b50]",
              "Sends extracted title + content to classification LLM. Returns:\n"
              "\u2022 slide_type_guess: title_slide | agenda_slide | dashboard_slide | "
              "comparison_slide | appendix_data_slide | chart_slide | content_slide\n"
              "\u2022 contains_chart, contains_dashboard, contains_comparison_layout, "
              "has_dense_visual_content (booleans)\n"
              "\u2022 confidence + rationale", "llm")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "7. Build PageResult [PROGRAMMATIC]",
              "PageResult with slide metadata in metadata dict")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "8. OUTPUT: ExtractionResult",
              "filetype='pptx', pages=[PageResult per slide], pages_failed=0", "output")
    r += 2

    _merged_cell(ws, r, c1, c2,
                 "KEY DIFFERENCE: No context passing between slides (each is independent). "
                 "Slide classification is a separate LLM call after extraction.",
                 FILLS["programmatic"], FONTS["small"])
    r += 2
    _legend(ws, r, c1)


# ── XLSX Sheet ─────────────────────────────────────────────────


def _build_xlsx_sheet(wb):
    ws = wb.create_sheet("XLSX Processor")
    _set_col_widths(ws, [4, 30, 30, 30, 4])
    c1, c2 = 2, 4
    r = 1

    _merged_cell(ws, r, c1, c2, "XLSX PROCESSOR FLOW", FILLS["title"], FONTS["title"])
    r += 1
    _merged_cell(ws, r, c1, c2,
                 "Input: .xlsx file  |  Output: ExtractionResult with PageResult per sheet",
                 FILLS["title"], FONTS["small"])
    r += 2

    r = _step(ws, r, c1, c2, "1. INPUT: Raw XLSX File",
              "file_path passed to process_xlsx()", "input")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "2. Open Workbook Twice [PROGRAMMATIC]",
              "Formula workbook (data_only=False): reads formula strings. "
              "Cached workbook (data_only=True): reads computed values. "
              "Both needed to detect formulas while showing correct cell values.")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "3. FOR EACH SHEET (sequential):",
              "Iterates workbook.sheetnames in order", "programmatic")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _decision(ws, r, c1, c2, "Is it a chartsheet?",
                  "Skip cell collection, mark as chartsheet",
                  "Build cached values, collect cells")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "4. Serialize Sheet [PROGRAMMATIC]",
              "Collects all populated cells, builds compact markdown table "
              "with row numbers and column letters. Computes metadata: "
              "row/column counts, formula count, merged ranges, used range.")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _decision(ws, r, c1, c2, "Does sheet have charts or images?",
                  "Describe each visual via LLM",
                  "Skip visual extraction")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "5. Visual Extraction [LLM \u2b50] (if visuals present)",
              "Charts: sends chart metadata (series, axis labels, data points) "
              "to LLM tool call for structured description.\n"
              "Images: sends raw image bytes to LLM vision for description.\n"
              "Each visual becomes a visual_region with bidirectional grid linking.", "llm")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "6. Region Detection [PROGRAMMATIC]",
              "Splits cells into contiguous regions via blank row/column bands. "
              "Scores each region with 5-component dense formula "
              "(shape 0.35, fill 0.25, row_bias 0.20, header 0.10, text_brevity 0.10). "
              "Native Excel tables always score 1.0. Threshold: \u2265 0.55 = dense candidate.")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _decision(ws, r, c1, c2, "content_kind?",
                  "empty \u2192 auto-classify as empty_sheet (no LLM)",
                  "visual_only \u2192 auto-classify as page_like (no LLM)")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "7. Sheet Classification [LLM \u2b50] (grid content only)",
              "Sends serialized sheet + region analysis + dense score to LLM. "
              "Returns: handling_mode (page_like | dense_table_candidate), "
              "confidence, rationale. Dense score is context, not the decision \u2014 "
              "LLM makes the final call.", "llm")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "8. Token Budget Check [PROGRAMMATIC]",
              "Counts tokens in serialized content. Sets threshold_exceeded flag "
              "if estimated_tokens > configured token_limit.")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "9. OUTPUT: ExtractionResult",
              "filetype='xlsx', one PageResult per sheet. Metadata includes: "
              "classification, dense_score, region details, visual descriptions, "
              "formula_cells count, token estimates.", "output")
    r += 2

    _merged_cell(ws, r, c1, c2,
                 "KEY DIFFERENCE: No vision rendering. Content extracted directly "
                 "from openpyxl cell values. Charts/images get separate LLM calls. "
                 "Classification is LLM-based but informed by programmatic dense score.",
                 FILLS["programmatic"], FONTS["small"])
    r += 2
    _legend(ws, r, c1)


# ── Vision Pipeline Sheet ──────────────────────────────────────


def _build_vision_sheet(wb):
    ws = wb.create_sheet("Vision Pipeline")
    _set_col_widths(ws, [4, 22, 22, 22, 22, 4])
    c1, c2 = 2, 5
    r = 1

    _merged_cell(ws, r, c1, c2, "SHARED VISION PIPELINE (process_page)", FILLS["title"], FONTS["title"])
    r += 1
    _merged_cell(ws, r, c1, c2,
                 "Used by PDF, DOCX, and PPTX processors. 4-tier fallback chain.",
                 FILLS["title"], FONTS["small"])
    r += 2

    r = _step(ws, r, c1, c2, "INPUT: PNG Image Bytes + Prompt",
              "Rendered page/slide image + extraction prompt (may include prior-page context)", "input")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "ATTEMPT 1: Full DPI, auto detail [LLM \u2b50]",
              "Sends full-resolution image with detail='auto'. "
              "Base64-encodes PNG, builds multimodal message, calls LLM with tool_choice=required. "
              "Retries 3x on transient errors (rate limit, timeout, connection, server error).", "llm")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _decision(ws, r, c1, c2, "Success?",
                  "Return PageResult (method='full_dpi')",
                  "Fall through to Attempt 2")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "ATTEMPT 2: High Detail [LLM \u2b50]",
              "Same image, explicit detail='high' parameter. "
              "Forces higher-resolution image tiles in the vision API.", "llm")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _decision(ws, r, c1, c2, "Success?",
                  "Return PageResult (method='high_detail')",
                  "Fall through to Attempt 3")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "ATTEMPT 3: Half Resolution [LLM \u2b50]",
              "shrink_image() halves resolution via fitz.Pixmap.shrink(1). "
              "Reduces token cost and may avoid content-length limits.", "llm")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _decision(ws, r, c1, c2, "Success?",
                  "Return PageResult (method='half_dpi')",
                  "Fall through to Attempt 4")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _step(ws, r, c1, c2, "ATTEMPT 4: Split Halves [LLM \u2b50 \u00d7 2]",
              "split_image() splits based on aspect ratio:\n"
              "\u2022 Portrait/square (w \u2264 h\u00d71.5): split top/bottom\n"
              "\u2022 Landscape (w > h\u00d71.5): split left/right\n"
              "First half extracted, then second half gets augmented prompt with "
              "first half's content as context. Results concatenated.", "llm")
    _arrow_row(ws, r, c1, c2); r += 1

    r = _decision(ws, r, c1, c2, "Success?",
                  "Return PageResult (method='split_halves')",
                  "RAISE RuntimeError \u2014 all attempts failed")
    r += 2

    r = _step(ws, r, c1, c2, "OUTPUT: PageResult",
              "page_number, page_title, content (markdown), "
              "method (which attempt succeeded), metadata={}", "output")
    r += 2

    _merged_cell(ws, r, c1, c2,
                 "ALL-OR-NOTHING: If all 4 attempts fail, RuntimeError propagates up. "
                 "The calling processor (PDF/DOCX/PPTX) aborts the entire file. "
                 "No partial extraction.",
                 FILLS["error"], FONTS["small"])
    r += 2
    _legend(ws, r, c1)


# ── Main ───────────────────────────────────────────────────────


def main():
    wb = Workbook()
    _build_pdf_sheet(wb.active)
    _build_docx_sheet(wb)
    _build_pptx_sheet(wb)
    _build_xlsx_sheet(wb)
    _build_vision_sheet(wb)

    for ws in wb.worksheets:
        ws.sheet_properties.pageSetUpPr = None

    wb.save(OUTPUT)
    print(f"Generated: {OUTPUT}")
    print(f"  Sheets: {', '.join(ws.title for ws in wb.worksheets)}")


if __name__ == "__main__":
    main()
