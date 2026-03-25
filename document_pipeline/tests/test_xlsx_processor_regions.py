"""Region-aware tests for XLSX processor classification."""

import json
from unittest.mock import MagicMock, patch

from openpyxl import Workbook
from openpyxl.worksheet.table import Table

from ingestion.processors.xlsx.processor import (
    _build_classifier_message,
    _build_region_metadata,
    _build_region_preview,
    _region_sample_text,
    _truncate_region_cell,
    process_xlsx,
)
from ingestion.processors.xlsx.layout import SheetRegion


def _make_fake_encoding() -> MagicMock:
    """Build a deterministic tokenizer stub for tests."""
    encoding = MagicMock()
    encoding.encode.side_effect = lambda text: [0] * max(1, len(text) // 8)
    return encoding


def _make_prompt() -> dict:
    """Build a minimal XLSX classification prompt for testing."""
    return {
        "stage": "xlsx_classification",
        "system_prompt": "Choose the worksheet handling mode.",
        "user_prompt": "Decide whether the sheet is page_like or dense.",
        "tool_choice": "required",
        "tools": [
            {
                "type": "function",
                "function": {
                    "name": "classify_xlsx_sheet",
                    "parameters": {},
                },
            }
        ],
    }


def _make_llm_response(
    handling_mode: str,
    confidence: float = 0.9,
    rationale: str = "classification rationale",
) -> dict:
    """Build a minimal tool-calling response payload."""
    return {
        "choices": [
            {
                "message": {
                    "tool_calls": [
                        {
                            "function": {
                                "arguments": json.dumps(
                                    {
                                        "handling_mode": handling_mode,
                                        "confidence": confidence,
                                        "rationale": rationale,
                                    }
                                )
                            }
                        }
                    ]
                }
            }
        ]
    }


def _save_workbook(path, sheets: list[tuple[str, list[list[object]]]]) -> None:
    """Create a workbook with the provided sheet data."""
    workbook = Workbook()
    first_sheet = workbook.active
    first_sheet.title = sheets[0][0]
    for row in sheets[0][1]:
        first_sheet.append(row)

    for title, rows in sheets[1:]:
        sheet = workbook.create_sheet(title=title)
        for row in rows:
            sheet.append(row)

    workbook.save(path)
    workbook.close()


def test_build_region_preview_returns_empty_without_region():
    """Missing dense candidates render as an empty region preview."""
    assert _build_region_preview(None) == ""


def test_truncate_region_cell_shortens_long_values():
    """Region preview cells are truncated to the preview limit."""
    truncated = _truncate_region_cell("x" * 90)

    assert truncated.endswith("...")
    assert len(truncated) == 80


def test_region_sample_text_returns_placeholder_for_empty_rows():
    """Regions with no populated values fall back to a placeholder sample."""
    region = SheetRegion(
        region_id="region_1",
        region_type="framing",
        min_row=1,
        min_col=1,
        max_row=2,
        max_col=2,
        non_empty_cells=0,
        row_count=2,
        column_count=2,
        row_numbers=[1, 2],
        column_numbers=[1, 2],
        rows=[{}, {}],
    )

    assert _region_sample_text(region) == "n/a"


def test_region_sample_text_truncates_long_values():
    """Long framing samples are shortened to the preview limit."""
    region = SheetRegion(
        region_id="region_1",
        region_type="framing",
        min_row=1,
        min_col=1,
        max_row=1,
        max_col=1,
        non_empty_cells=1,
        row_count=1,
        column_count=1,
        row_numbers=[1],
        column_numbers=[1],
        rows=[{1: "x" * 100}],
    )

    assert _region_sample_text(region) == "x" * 77 + "..."


def test_build_region_metadata_returns_defaults_for_blank_sheet():
    """Blank worksheets return empty region metadata."""
    workbook = Workbook()
    sheet = workbook.active

    metadata = _build_region_metadata(sheet, {})

    assert metadata["region_count"] == 0
    assert metadata["sheet_regions"] == []
    assert metadata["dense_table_region"] is None
    assert metadata["framing_summary"] == ""
    workbook.close()


def test_build_classifier_message_uses_sheet_preview_without_candidate():
    """Sheets without a selected dense region keep the whole-sheet preview."""
    message = _build_classifier_message(
        prompt=_make_prompt(),
        sheet_content=(
            "# Sheet: Notes\n| Row | A |\n| --- | --- |\n| 1 | Note |"
        ),
        sheet_metadata={
            "sheet_name": "Notes",
            "sheet_kind": "worksheet",
            "used_range": "A1",
            "row_count": 1,
            "column_count": 1,
            "row_span": 1,
            "column_span": 1,
            "blank_rows_omitted": 0,
            "blank_columns_omitted": 0,
            "merged_range_count": 0,
            "chart_count": 0,
            "image_count": 0,
            "estimated_tokens": 12,
            "token_limit": 12000,
            "region_count": 1,
            "sheet_regions": [
                {
                    "region_id": "region_1",
                    "region_type": "framing",
                    "used_range": "A1",
                    "row_count": 1,
                    "column_count": 1,
                    "dense_score": 0.0,
                }
            ],
            "dense_table_candidate_detected": False,
            "dense_table_used_range": "",
            "dense_table_region_preview": "",
            "framing_summary": (
                "- region_1: framing, range=A1, rows=1, columns=1, "
                "sample=Note"
            ),
        },
    )

    assert "## Local layout analysis" in message
    assert "## Framing regions" in message
    assert "## Sheet preview" in message
    assert "## Candidate dense-table region" not in message


@patch(
    "ingestion.processors.xlsx.processor.get_stage_model_config",
    return_value={
        "model": "gpt-5-mini",
        "max_tokens": 800,
        "temperature": None,
    },
)
@patch(
    "ingestion.processors.xlsx.processor.tiktoken.encoding_for_model",
    return_value=_make_fake_encoding(),
)
@patch(
    "ingestion.processors.xlsx.processor.get_xlsx_sheet_token_limit",
    return_value=12000,
)
@patch(
    "ingestion.processors.xlsx.processor.load_prompt",
    return_value=_make_prompt(),
)
def test_process_xlsx_mixed_layout_sheet_records_dense_table_region(
    _prompt, _token_limit, _encoding, _model_config, tmp_path
):
    """Mixed-layout sheets preserve dense-region bounds and framing notes."""
    file_path = tmp_path / "mixed_layout.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Exposure"
    sheet["A1"] = "Portfolio Summary"
    sheet["B1"] = "Prepared by Finance"
    sheet["A3"] = "Account"
    sheet["B3"] = "Region"
    sheet["C3"] = "Amount"
    sheet["A4"] = "A-1"
    sheet["B4"] = "East"
    sheet["C4"] = 100
    sheet["A5"] = "A-2"
    sheet["B5"] = "West"
    sheet["C5"] = 200
    sheet["A6"] = "A-3"
    sheet["B6"] = "North"
    sheet["C6"] = 300
    sheet.add_table(Table(displayName="ExposureTable", ref="A3:C6"))
    workbook.save(file_path)
    workbook.close()

    llm = MagicMock()
    llm.call.return_value = _make_llm_response(
        "page_like", 0.9, "Framing indicates report-style content."
    )

    result = process_xlsx(str(file_path), llm)

    page = result.pages[0]
    assert page.metadata["region_count"] == 2
    assert page.metadata["dense_table_candidate_detected"] is True
    assert page.metadata["dense_table_used_range"] == "A3:C6"
    assert page.metadata["dense_table_region"]["used_range"] == "A3:C6"
    assert page.metadata["dense_table_region"]["native_table_name"] == (
        "ExposureTable"
    )
    assert page.metadata["dense_table_region"]["metadata"]["table_ref"] == (
        "A3:C6"
    )
    assert page.metadata["framing_regions"] == [
        {
            "region_id": "region_1",
            "region_type": "framing",
            "used_range": "A1:B1",
            "row_count": 1,
            "column_count": 2,
            "non_empty_cells": 2,
            "dense_score": 0.0,
            "native_table_name": "",
        }
    ]


@patch(
    "ingestion.processors.xlsx.processor.get_stage_model_config",
    return_value={
        "model": "gpt-5-mini",
        "max_tokens": 800,
        "temperature": None,
    },
)
@patch(
    "ingestion.processors.xlsx.processor.tiktoken.encoding_for_model",
    return_value=_make_fake_encoding(),
)
@patch(
    "ingestion.processors.xlsx.processor.get_xlsx_sheet_token_limit",
    return_value=12000,
)
@patch(
    "ingestion.processors.xlsx.processor.load_prompt",
    return_value=_make_prompt(),
)
def test_process_xlsx_classifier_prompt_uses_region_preview(
    _prompt, _token_limit, _encoding, _model_config, tmp_path
):
    """Local dense regions drive the classifier prompt."""
    file_path = tmp_path / "region_prompt.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Exposure"
    sheet["A1"] = "Portfolio Summary"
    sheet["B1"] = "Prepared by Finance"
    sheet["A3"] = "Account"
    sheet["B3"] = "Region"
    sheet["C3"] = "Amount"
    sheet["A4"] = "A-1"
    sheet["B4"] = "East"
    sheet["C4"] = 100
    sheet["A5"] = "A-2"
    sheet["B5"] = "West"
    sheet["C5"] = 200
    sheet["A6"] = "A-3"
    sheet["B6"] = "North"
    sheet["C6"] = "x" * 100
    sheet.add_table(Table(displayName="ExposureTable", ref="A3:C6"))
    workbook.save(file_path)
    workbook.close()

    llm = MagicMock()
    llm.call.return_value = _make_llm_response(
        "page_like", 0.9, "Framing indicates report-style content."
    )

    process_xlsx(str(file_path), llm)

    prompt_text = llm.call.call_args.kwargs["messages"][1]["content"]
    assert "## Local layout analysis" in prompt_text
    assert "## Candidate dense-table region" in prompt_text
    assert "## Framing outside candidate region" in prompt_text
    assert "- Selected dense-table region: A3:C6" in prompt_text
    assert (
        "... omitted candidate region rows or columns ..." not in prompt_text
    )
    assert "## Sheet preview" not in prompt_text
    assert "x" * 77 + "..." in prompt_text


@patch(
    "ingestion.processors.xlsx.processor.get_stage_model_config",
    return_value={
        "model": "gpt-5-mini",
        "max_tokens": 800,
        "temperature": None,
    },
)
@patch(
    "ingestion.processors.xlsx.processor.tiktoken.encoding_for_model",
    return_value=_make_fake_encoding(),
)
@patch(
    "ingestion.processors.xlsx.processor.get_xlsx_sheet_token_limit",
    return_value=12000,
)
@patch(
    "ingestion.processors.xlsx.processor.load_prompt",
    return_value=_make_prompt(),
)
def test_process_xlsx_sheet_without_dense_region_has_empty_dense_metadata(
    _prompt, _token_limit, _encoding, _model_config, tmp_path
):
    """Low-density sheets keep framing metadata and use full-sheet preview."""
    file_path = tmp_path / "notes.xlsx"
    _save_workbook(file_path, [("Notes", [["Owner", "Treasury"]])])
    llm = MagicMock()
    llm.call.return_value = _make_llm_response(
        "page_like", 0.92, "Single note row is page-like."
    )

    result = process_xlsx(str(file_path), llm)

    page = result.pages[0]
    prompt_text = llm.call.call_args.kwargs["messages"][1]["content"]
    assert page.metadata["region_count"] == 1
    assert page.metadata["dense_table_candidate_detected"] is False
    assert page.metadata["dense_table_region"] is None
    assert page.metadata["dense_table_used_range"] == ""
    assert page.metadata["framing_regions"] == [
        {
            "region_id": "region_1",
            "region_type": "framing",
            "used_range": "A1:B1",
            "row_count": 1,
            "column_count": 2,
            "non_empty_cells": 2,
            "dense_score": 0.0,
            "native_table_name": "",
        }
    ]
    assert "## Framing regions" in prompt_text
    assert "## Sheet preview" in prompt_text
