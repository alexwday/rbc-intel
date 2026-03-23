"""Tests for XLSX processor classification."""

import base64
from copy import deepcopy
from datetime import date, datetime
import json
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as XLImage

from ingestion.processors.xlsx import (
    _FallbackEncoding,
    _FallbackTiktoken,
    _append_visual_descriptions,
    _build_chart_context,
    _build_region_metadata,
    _build_cached_values,
    _build_preview,
    _build_used_range,
    _call_visual_prompt_with_retry,
    _classify_sheet_with_retry,
    _describe_image_visual,
    _escape_table_text,
    _extract_anchor_position,
    _extract_chart_title,
    _extract_series_name,
    _extract_series_points,
    _extract_title_text,
    _get_encoder,
    _load_tiktoken,
    _load_reference_values,
    _normalize_cell_value,
    _open_workbooks,
    _serialize_sheet,
    _visual_overlaps_region,
    process_xlsx,
)
from ingestion.utils.xlsx_layout import SheetRegion


def _make_fake_encoding() -> MagicMock:
    """Build a deterministic tokenizer stub for tests."""
    encoding = MagicMock()
    encoding.encode.side_effect = lambda text: [0] * max(1, len(text) // 8)
    return encoding


def _make_prompt() -> dict:
    """Build a minimal XLSX classification prompt for testing."""
    return {
        "stage": "xlsx_classification",
        "system_prompt": "Choose the worksheet handling mode.",
        "user_prompt": "Decide whether the sheet is page_like or dense.",
        "tool_choice": "required",
        "tools": [
            {
                "type": "function",
                "function": {
                    "name": "classify_xlsx_sheet",
                    "parameters": {},
                },
            }
        ],
    }


def _make_visual_prompt() -> dict:
    """Build a minimal visual-description prompt for XLSX charts/images."""
    return {
        "stage": "extraction",
        "system_prompt": "Describe the worksheet visual.",
        "user_prompt": "Extract the worksheet visual content.",
        "tool_choice": "required",
        "tools": [
            {
                "type": "function",
                "function": {
                    "name": "extract_page_content",
                    "parameters": {},
                },
            }
        ],
    }


def _make_prompt_loader(name: str) -> dict:
    """Return the right prompt fixture for the requested XLSX prompt name."""
    if name == "xlsx_sheet_classification":
        return _make_prompt()
    if name == "xlsx_visual_extraction_vision":
        return _make_visual_prompt()
    raise AssertionError(f"Unexpected prompt requested: {name}")


def _make_llm_response(
    handling_mode: str,
    confidence: float = 0.9,
    rationale: str = "classification rationale",
) -> dict:
    """Build a minimal tool-calling response payload."""
    return {
        "choices": [
            {
                "message": {
                    "tool_calls": [
                        {
                            "function": {
                                "arguments": json.dumps(
                                    {
                                        "handling_mode": handling_mode,
                                        "confidence": confidence,
                                        "rationale": rationale,
                                    }
                                )
                            }
                        }
                    ]
                }
            }
        ]
    }


def _make_visual_response(
    page_title: str = "Revenue Trend",
    content: str = (
        "### Charts and Visuals\n"
        "**Revenue Trend**\n"
        "- Type: bar\n"
        "- Data points: Jan 10, Feb 20"
    ),
) -> dict:
    """Build a tool-calling response payload for visual descriptions."""
    return {
        "choices": [
            {
                "message": {
                    "tool_calls": [
                        {
                            "function": {
                                "arguments": json.dumps(
                                    {
                                        "page_title": page_title,
                                        "content": content,
                                    }
                                )
                            }
                        }
                    ]
                }
            }
        ]
    }


def _save_workbook(path, sheets: list[tuple[str, list[list[object]]]]) -> None:
    """Create a workbook with the provided sheet data."""
    workbook = Workbook()
    first_sheet = workbook.active
    first_sheet.title = sheets[0][0]
    for row in sheets[0][1]:
        first_sheet.append(row)

    for title, rows in sheets[1:]:
        sheet = workbook.create_sheet(title=title)
        for row in rows:
            sheet.append(row)

    workbook.save(path)
    workbook.close()


def _save_workbook_with_visual_tabs(path) -> None:
    """Create a workbook with a chart-only worksheet and chartsheet."""
    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "Data"
    data_sheet.append(["Month", "Value"])
    data_sheet.append(["Jan", 10])
    data_sheet.append(["Feb", 20])

    dashboard = workbook.create_sheet(title="Dashboard")
    dashboard_chart = BarChart()
    dashboard_chart.title = "Revenue"
    data_range = Reference(data_sheet, min_col=2, min_row=1, max_row=3)
    category_range = Reference(data_sheet, min_col=1, min_row=2, max_row=3)
    dashboard_chart.add_data(data_range, titles_from_data=True)
    dashboard_chart.set_categories(category_range)
    dashboard.add_chart(dashboard_chart, "A1")

    chart_tab = workbook.create_chartsheet(title="ChartTab")
    chart_tab.add_chart(deepcopy(dashboard_chart))

    workbook.save(path)
    workbook.close()


def _save_workbook_with_image_tab(path, image_path) -> None:
    """Create a workbook with an image-only worksheet."""
    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "Data"
    data_sheet.append(["Metric", "Value"])
    data_sheet.append(["Revenue", 20])

    dashboard = workbook.create_sheet(title="Dashboard")
    dashboard.add_image(XLImage(str(image_path)), "B2")

    workbook.save(path)
    workbook.close()


def test_normalize_cell_value_handles_empty_and_bool():
    """Normalizes empty cells and booleans deterministically."""
    assert _normalize_cell_value(None) == ""
    assert _normalize_cell_value(True) == "TRUE"
    assert _normalize_cell_value(False) == "FALSE"


def test_normalize_cell_value_normalizes_carriage_returns():
    """Carriage returns are converted to plain newlines."""
    assert _normalize_cell_value("line1\r\nline2") == "line1\nline2"
    assert _normalize_cell_value("line1\rline2") == "line1\nline2"


def test_normalize_cell_value_formats_datetime_with_time():
    """Datetimes with a time component include hours and minutes."""
    assert (
        _normalize_cell_value(datetime(2026, 3, 17, 14, 30, 0))
        == "2026-03-17 14:30:00"
    )


def test_normalize_cell_value_formats_midnight_datetime_as_date():
    """Midnight datetimes render as date-only strings."""
    assert _normalize_cell_value(datetime(2026, 3, 17)) == "2026-03-17"


def test_normalize_cell_value_formats_date():
    """Pure date objects render as YYYY-MM-DD."""
    assert _normalize_cell_value(date(2026, 3, 17)) == "2026-03-17"


def test_escape_table_text_escapes_pipes_and_newlines():
    """Pipes and newlines are escaped for markdown table safety."""
    assert _escape_table_text("a|b") == "a\\|b"
    assert _escape_table_text("line1\nline2") == "line1<br>line2"
    assert _escape_table_text("a|b\nc|d") == "a\\|b<br>c\\|d"


def test_build_preview_returns_short_content_unchanged():
    """Content within the head+tail threshold is returned as-is."""
    content = "\n".join(f"row {i}" for i in range(30))
    assert _build_preview(content) == content


def test_build_preview_truncates_long_content():
    """Content exceeding the threshold is spliced with an omission marker."""
    lines = [f"row {i}" for i in range(60)]
    preview = _build_preview("\n".join(lines))
    preview_lines = preview.splitlines()

    assert preview_lines[:24] == lines[:24]
    assert preview_lines[24] == "... omitted sheet rows ..."
    assert preview_lines[25:] == lines[-10:]


def test_build_preview_truncates_wide_lines():
    """Lines exceeding the width limit are truncated with an ellipsis."""
    wide_line = "x" * 500
    short_line = "short"
    content = "\n".join([short_line, wide_line])
    preview = _build_preview(content)
    preview_lines = preview.splitlines()

    assert len(preview_lines) == 2
    assert preview_lines[0] == short_line
    assert len(preview_lines[1]) == 304
    assert preview_lines[1].endswith(" ...")


def test_build_used_range_single_cell():
    """Single-cell bounds stay in simple A1 notation."""
    assert _build_used_range((1, 1, 1, 1)) == "A1"


@patch(
    "ingestion.processors.xlsx.tiktoken.get_encoding",
    return_value=_make_fake_encoding(),
)
@patch(
    "ingestion.processors.xlsx.tiktoken.encoding_for_model",
    side_effect=KeyError("missing model mapping"),
)
def test_get_encoder_falls_back_to_o200k_base(
    _encoding_for_model, _get_encoding
):
    """Unknown model names fall back to the base tokenizer."""
    encoding = _get_encoder("custom-model")
    assert encoding.encode("hello")


def test_fallback_tiktoken_provides_minimal_encoder():
    """Fallback tokenizer returns coarse encoders for local counting."""
    tokenizer = _FallbackTiktoken()

    assert isinstance(
        tokenizer.encoding_for_model("gpt-5-mini"), _FallbackEncoding
    )
    assert isinstance(tokenizer.get_encoding("o200k_base"), _FallbackEncoding)
    assert tokenizer.encoding_for_model("gpt-5-mini").encode("") == [0]
    text_20_chars = "a" * 20
    tokens = tokenizer.encoding_for_model("gpt-5-mini").encode(text_20_chars)
    assert len(tokens) == 5


@patch("ingestion.processors.xlsx.import_module", side_effect=ImportError)
def test_load_tiktoken_falls_back_on_import_error(_import_module):
    """Import errors fall back to the local tokenizer shim."""
    tokenizer = _load_tiktoken()

    assert isinstance(tokenizer, _FallbackTiktoken)


def test_fallback_encoding_decode_raises_not_implemented():
    """Fallback tokenizer does not support decode operations."""
    with pytest.raises(NotImplementedError, match="does not support decode"):
        _FallbackEncoding().decode([0, 1, 2])


def test_extract_chart_title_returns_empty_without_title():
    """Missing chart titles serialize as empty text."""
    assert _extract_chart_title(SimpleNamespace()) == ""


def test_extract_title_text_strips_plain_strings():
    """String titles are normalized directly without rich-text parsing."""
    assert _extract_title_text(" Revenue ") == "Revenue"


def test_extract_chart_title_reads_field_text():
    """Chart titles can be recovered from field text runs."""
    chart = SimpleNamespace(
        title=SimpleNamespace(
            tx=SimpleNamespace(
                rich=SimpleNamespace(
                    p=[
                        SimpleNamespace(
                            r=[],
                            fld=[SimpleNamespace(t="Field Title")],
                        )
                    ]
                )
            )
        )
    )

    assert _extract_chart_title(chart) == "Field Title"


def test_extract_anchor_position_handles_string_and_from_marker():
    """Anchors from strings and from_ markers serialize consistently."""
    from_marker = SimpleNamespace(from_=SimpleNamespace(row=1, col=2))

    assert _extract_anchor_position("B3:C4") == {
        "anchor_row": 3,
        "anchor_column": 2,
        "anchor_cell": "B3",
    }
    assert _extract_anchor_position(from_marker) == {
        "anchor_row": 2,
        "anchor_column": 3,
        "anchor_cell": "C2",
    }


def test_load_reference_values_returns_empty_on_invalid_or_missing_sheet():
    """Invalid Excel refs and missing sheets fail closed."""
    workbook = Workbook()
    cached_workbook = Workbook()

    assert not _load_reference_values(
        workbook,
        cached_workbook,
        "not-a-ref",
    )
    assert not _load_reference_values(
        workbook,
        cached_workbook,
        "'Missing'!A1:A2",
    )

    workbook.close()
    cached_workbook.close()


def test_extract_series_name_handles_literal_and_fallback_values():
    """Series names prefer literals and otherwise use fallback values."""
    workbook = Workbook()
    cached_workbook = Workbook()
    literal_series = SimpleNamespace(tx=SimpleNamespace(v="Revenue"))
    fallback_series = SimpleNamespace(tx=None)

    assert (
        _extract_series_name(literal_series, workbook, cached_workbook, 1)
        == "Revenue"
    )
    assert (
        _extract_series_name(fallback_series, workbook, cached_workbook, 2)
        == "Series 2"
    )

    workbook.close()
    cached_workbook.close()


def test_extract_series_points_caps_output_to_point_limit():
    """Large chart series are truncated to the configured point limit."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Category", "Value"])
    for index in range(1, 20):
        sheet.append([f"C{index}", index])
    cached_workbook = workbook
    series = SimpleNamespace(
        cat=SimpleNamespace(numRef=SimpleNamespace(f="'Data'!$A$2:$A$20")),
        val=SimpleNamespace(numRef=SimpleNamespace(f="'Data'!$B$2:$B$20")),
    )

    points = _extract_series_points(series, workbook, cached_workbook)

    assert len(points) == 12
    assert points[0] == "C1: 1"
    workbook.close()


def test_extract_series_points_skips_blank_category_value_pairs():
    """Blank category/value pairs are skipped."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Data"
    sheet.append(["Category", "Value"])
    sheet.append([None, None])
    sheet.append(["North", 15])
    cached_workbook = workbook
    series = SimpleNamespace(
        cat=SimpleNamespace(numRef=SimpleNamespace(f="'Data'!$A$2:$A$3")),
        val=SimpleNamespace(numRef=SimpleNamespace(f="'Data'!$B$2:$B$3")),
    )

    points = _extract_series_points(series, workbook, cached_workbook)

    assert points == ["North: 15"]
    workbook.close()


def test_build_chart_context_reports_when_points_are_unavailable():
    """Chart context notes when workbook refs cannot recover points."""
    workbook = Workbook()
    series = SimpleNamespace(
        tx=SimpleNamespace(v="Revenue"),
        cat=None,
        val=None,
    )
    chart = SimpleNamespace(
        anchor="B2",
        title=None,
        x_axis=SimpleNamespace(title=None),
        y_axis=SimpleNamespace(title=None),
        ser=[series],
    )

    context = _build_chart_context(
        "Dashboard",
        "visual_region_1",
        chart,
        workbook,
        workbook,
    )

    assert "- Data points: none recovered from workbook refs" in context
    workbook.close()


@patch("ingestion.processors.xlsx.time.sleep")
def test_call_visual_prompt_with_retry_raises_after_exhaustion(mock_sleep):
    """Visual description retries eventually fail the workbook cleanly."""
    llm = MagicMock()
    llm.call.side_effect = RuntimeError("visual failure")
    visual_label = "chart 'Sheet:visual_region_1'"

    with pytest.raises(RuntimeError, match="visual failure"):
        _call_visual_prompt_with_retry(
            llm,
            _make_visual_prompt(),
            "chart context",
            visual_label,
        )
    assert mock_sleep.call_count == 2


@patch(
    "ingestion.processors.xlsx.get_xlsx_classification_max_retries",
    return_value=0,
)
def test_call_visual_prompt_with_retry_raises_when_retries_disabled(_retries):
    """A zero-retry configuration fails without calling the LLM."""
    llm = MagicMock()

    with pytest.raises(RuntimeError, match="exited retry loop"):
        _call_visual_prompt_with_retry(
            llm=llm,
            prompt=_make_visual_prompt(),
            user_context="chart context",
            visual_label="chart 'Sheet:visual_region_1'",
        )
    llm.call.assert_not_called()


def test_describe_image_visual_requires_image_bytes():
    """Images without a byte loader fail fast."""
    with pytest.raises(RuntimeError, match="Embedded image bytes"):
        _describe_image_visual(
            llm=MagicMock(),
            prompt=_make_visual_prompt(),
            sheet_name="Dashboard",
            image=SimpleNamespace(anchor="B2"),
            visual_index=1,
        )


def test_visual_overlaps_region_rejects_missing_anchor_coordinates():
    """Visual overlap checks fail closed when anchors are incomplete."""
    region = SheetRegion(
        region_id="region_1",
        region_type="framing",
        min_row=1,
        min_col=1,
        max_row=3,
        max_col=3,
        non_empty_cells=4,
        row_count=3,
        column_count=3,
        row_numbers=[1, 2, 3],
        column_numbers=[1, 2, 3],
        rows=[{1: "A"}],
    )

    assert (
        _visual_overlaps_region(
            {"anchor_row": "1", "anchor_column": None},
            region,
        )
        is False
    )


def test_build_region_metadata_marks_mixed_and_small_table_regions():
    """Region metadata classifies overlapping visuals and small tables."""
    workbook = Workbook()
    mixed_sheet = workbook.active
    mixed_sheet.title = "Mixed"
    mixed_sheet.append(["Month", "Value"])
    mixed_sheet.append(["Jan", 10])
    mixed_sheet.append(["Feb", 20])

    metadata = _build_region_metadata(
        mixed_sheet,
        {},
        [
            {
                "region_id": "visual_region_1",
                "anchor_row": 2,
                "anchor_column": 2,
                "visual_kind": "chart",
            }
        ],
    )

    assert metadata["mixed_regions"][0]["region_id"] == "region_1"
    assert metadata["dense_table_regions"][0]["typed_region_type"] == "mixed"

    small_sheet = workbook.create_sheet(title="Small")
    long_text = (
        "Quarter-end snapshot commentary with narrative detail that reads "
        "like framing instead of structured row data."
    )
    small_sheet.append([long_text, long_text, long_text, long_text])
    small_sheet.append([long_text])

    small_metadata = _build_region_metadata(small_sheet, {}, [])

    assert len(small_metadata["small_table_regions"]) == 1
    workbook.close()


def test_append_visual_descriptions_includes_linked_grid_regions():
    """Visual description rendering includes linked grid-region metadata."""
    content = _append_visual_descriptions(
        "# Sheet: Dashboard",
        [
            {
                "region_id": "visual_region_1",
                "visual_kind": "chart",
                "visual_title": "Revenue",
                "anchor_cell": "B2",
                "linked_grid_region_ids": ["region_1"],
                "description_content": "### Charts and Visuals\nRevenue",
            }
        ],
    )

    assert "Linked grid regions: region_1" in content


@patch(
    "ingestion.processors.xlsx.get_stage_model_config",
    return_value={
        "model": "gpt-5-mini",
        "max_tokens": 800,
        "temperature": None,
    },
)
@patch(
    "ingestion.processors.xlsx.tiktoken.encoding_for_model",
    return_value=_make_fake_encoding(),
)
@patch(
    "ingestion.processors.xlsx.get_xlsx_sheet_token_limit", return_value=12000
)
@patch("ingestion.processors.xlsx.load_prompt", return_value=_make_prompt())
def test_process_xlsx_small_sheet_page_like(
    _prompt, _token_limit, _encoding, _model_config, tmp_path
):
    """Small worksheet stays page-like and records metadata."""
    file_path = tmp_path / "small.xlsx"
    _save_workbook(
        file_path,
        [
            (
                "Summary",
                [
                    ["Metric", "Value"],
                    ["Revenue", 120],
                    ["Expenses", 90],
                    ["Margin", "=B2-B3"],
                ],
            )
        ],
    )
    llm = MagicMock()
    llm.call.return_value = _make_llm_response(
        "page_like", 0.88, "Compact summary"
    )

    result = process_xlsx(str(file_path), llm)

    assert result.filetype == "xlsx"
    assert result.total_pages == 1
    assert result.pages_succeeded == 1
    page = result.pages[0]
    assert page.page_title == "Summary"
    assert page.method == "xlsx_sheet_classification"
    assert page.metadata["classification"] == "page_like"
    assert page.metadata["contains_dense_table"] is False
    assert page.metadata["threshold_exceeded"] is False
    assert page.metadata["estimated_tokens"] > 0
    assert "Revenue" in page.content
    llm.call.assert_called_once()
    assert llm.call.call_args.kwargs["stage"] == "xlsx_classification"


@patch(
    "ingestion.processors.xlsx.get_stage_model_config",
    return_value={
        "model": "gpt-5-mini",
        "max_tokens": 800,
        "temperature": None,
    },
)
@patch(
    "ingestion.processors.xlsx.tiktoken.encoding_for_model",
    return_value=_make_fake_encoding(),
)
@patch("ingestion.processors.xlsx.get_xlsx_sheet_token_limit", return_value=80)
@patch("ingestion.processors.xlsx.load_prompt", return_value=_make_prompt())
def test_process_xlsx_dense_sheet_marks_threshold_and_dense_table(
    _prompt, _token_limit, _encoding, _model_config, tmp_path
):
    """Large worksheet records both token overflow and dense-table output."""
    file_path = tmp_path / "dense.xlsx"
    dense_rows = [["Account", "Region", "Quarter", "Amount", "Status"]]
    for index in range(1, 90):
        dense_rows.append(
            [
                f"Account {index}",
                f"Region {index % 5}",
                f"Q{(index % 4) + 1}",
                index * 1000,
                "Open" if index % 2 else "Closed",
            ]
        )
    _save_workbook(file_path, [("Transactions", dense_rows)])
    llm = MagicMock()
    llm.call.return_value = _make_llm_response(
        "dense_table_candidate",
        0.97,
        "Repeated records indicate a dense data grid.",
    )

    result = process_xlsx(str(file_path), llm)

    page = result.pages[0]
    assert page.metadata["contains_dense_table"] is True
    assert page.metadata["classification"] == "dense_table_candidate"
    assert page.metadata["threshold_exceeded"] is True
    assert page.metadata["estimated_tokens"] > page.metadata["token_limit"]
    assert page.metadata["region_count"] == 1
    assert page.metadata["dense_table_candidate_detected"] is True
    assert page.metadata["framing_summary"] == (
        "No separate framing regions detected."
    )


@patch(
    "ingestion.processors.xlsx.get_stage_model_config",
    return_value={
        "model": "gpt-5-mini",
        "max_tokens": 800,
        "temperature": None,
    },
)
@patch(
    "ingestion.processors.xlsx.tiktoken.encoding_for_model",
    return_value=_make_fake_encoding(),
)
@patch(
    "ingestion.processors.xlsx.get_xlsx_sheet_token_limit", return_value=12000
)
@patch("ingestion.processors.xlsx.load_prompt", return_value=_make_prompt())
def test_process_xlsx_preserves_multiple_dense_regions(
    _prompt, _token_limit, _encoding, _model_config, tmp_path
):
    """Sheets with two real tables retain both dense regions in metadata."""
    file_path = tmp_path / "two_tables.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Exports"
    for row in (
        ["Account", "Amount"],
        ["A-1", 100],
        ["A-2", 200],
        [None, None],
        ["Region", "Balance"],
        ["East", 300],
        ["West", 400],
    ):
        sheet.append(row)
    workbook.save(file_path)
    workbook.close()

    llm = MagicMock()
    llm.call.return_value = _make_llm_response(
        "dense_table_candidate",
        0.95,
        "The sheet contains multiple dense exports.",
    )

    result = process_xlsx(str(file_path), llm)

    page = result.pages[0]
    assert len(page.metadata["dense_table_regions"]) == 2
    assert {
        region["used_range"] for region in page.metadata["dense_table_regions"]
    } == {
        "A1:B3",
        "A5:B7",
    }
    assert page.metadata["dense_table_region"]["used_range"] in {
        "A1:B3",
        "A5:B7",
    }


@patch(
    "ingestion.processors.xlsx.get_stage_model_config",
    return_value={
        "model": "gpt-5-mini",
        "max_tokens": 800,
        "temperature": None,
    },
)
@patch(
    "ingestion.processors.xlsx.tiktoken.encoding_for_model",
    return_value=_make_fake_encoding(),
)
@patch(
    "ingestion.processors.xlsx.get_xlsx_sheet_token_limit", return_value=12000
)
@patch("ingestion.processors.xlsx.load_prompt", return_value=_make_prompt())
def test_process_xlsx_sparse_sheet_omits_blank_rows_and_columns(
    _prompt, _token_limit, _encoding, _model_config, tmp_path
):
    """Sparse worksheets serialize only populated rows and columns."""
    file_path = tmp_path / "sparse.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sparse"
    sheet["A1"] = "Top"
    sheet["Z100"] = "Bottom"
    workbook.save(file_path)
    workbook.close()

    llm = MagicMock()
    llm.call.return_value = _make_llm_response(
        "page_like", 0.85, "Sparse page."
    )

    result = process_xlsx(str(file_path), llm)

    page = result.pages[0]
    assert page.metadata["used_range"] == "A1:Z100"
    assert page.metadata["row_count"] == 2
    assert page.metadata["column_count"] == 2
    assert page.metadata["row_span"] == 100
    assert page.metadata["column_span"] == 26
    assert page.metadata["blank_rows_omitted"] == 98
    assert page.metadata["blank_columns_omitted"] == 24
    assert "| Row | A | Z |" in page.content
    assert "| 2 |" not in page.content


@patch(
    "ingestion.processors.xlsx.get_stage_model_config",
    return_value={
        "model": "gpt-5-mini",
        "max_tokens": 800,
        "temperature": None,
    },
)
@patch(
    "ingestion.processors.xlsx.tiktoken.encoding_for_model",
    return_value=_make_fake_encoding(),
)
@patch(
    "ingestion.processors.xlsx.get_xlsx_sheet_token_limit", return_value=12000
)
@patch(
    "ingestion.processors.xlsx.load_prompt",
    side_effect=_make_prompt_loader,
)
def test_process_xlsx_visual_only_tabs_are_not_empty(
    _prompt, _token_limit, _encoding, _model_config, tmp_path
):
    """Visual-only worksheets and chartsheets stay queryable."""
    file_path = tmp_path / "visual_tabs.xlsx"
    _save_workbook_with_visual_tabs(file_path)
    llm = MagicMock()
    llm.call.side_effect = [
        _make_llm_response("page_like", 0.9, "Compact grid."),
        _make_visual_response(
            page_title="Revenue",
            content="### Charts and Visuals\n**Revenue**\n- Type: bar",
        ),
        _make_visual_response(
            page_title="Revenue",
            content="### Charts and Visuals\n**Revenue**\n- Type: bar",
        ),
    ]

    result = process_xlsx(str(file_path), llm)

    assert result.total_pages == 3
    assert [page.page_title for page in result.pages] == [
        "Data",
        "Dashboard",
        "ChartTab",
    ]
    dashboard = result.pages[1]
    chart_tab = result.pages[2]
    assert dashboard.content != "## Empty Sheet"
    assert dashboard.metadata["content_kind"] == "visual_only"
    assert dashboard.metadata["sheet_layout_kind"] == "visual_only"
    assert dashboard.metadata["classification"] == "page_like"
    assert dashboard.metadata["chart_count"] == 1
    assert dashboard.metadata["region_count"] == 0
    assert dashboard.metadata["visual_region_count"] == 1
    assert dashboard.metadata["visual_regions"][0]["visual_kind"] == "chart"
    assert "## Visual Elements" in dashboard.content
    assert "## Visual Region Descriptions" in dashboard.content
    assert chart_tab.metadata["sheet_kind"] == "chartsheet"
    assert chart_tab.metadata["classification"] == "page_like"
    assert chart_tab.metadata["visual_region_count"] == 1
    assert "## Visual Region Descriptions" in chart_tab.content
    assert llm.call.call_count == 3


@patch(
    "ingestion.processors.xlsx.get_stage_model_config",
    return_value={
        "model": "gpt-5-mini",
        "max_tokens": 800,
        "temperature": None,
    },
)
@patch(
    "ingestion.processors.xlsx.tiktoken.encoding_for_model",
    return_value=_make_fake_encoding(),
)
@patch(
    "ingestion.processors.xlsx.get_xlsx_sheet_token_limit", return_value=12000
)
@patch(
    "ingestion.processors.xlsx.load_prompt",
    side_effect=_make_prompt_loader,
)
@patch("ingestion.processors.xlsx.call_vision")
def test_process_xlsx_visual_only_image_tab_uses_vision_description(
    mock_call_vision,
    _prompt,
    _token_limit,
    _encoding,
    _model_config,
    tmp_path,
):
    """Image-only tabs create first-class visual regions via vision."""
    image_path = tmp_path / "dash.png"
    image_path.write_bytes(
        base64.b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lE"
            "QVR42mP8/x8AAusB9Wl8tWQAAAAASUVORK5CYII="
        )
    )
    file_path = tmp_path / "image_tab.xlsx"
    _save_workbook_with_image_tab(file_path, image_path)
    mock_call_vision.return_value = (
        "Dashboard Snapshot",
        "### Charts and Visuals\n**Dashboard Snapshot**\n- Type: other",
    )
    llm = MagicMock()
    llm.call.return_value = _make_llm_response(
        "page_like", 0.9, "Compact grid."
    )

    result = process_xlsx(str(file_path), llm)

    dashboard = result.pages[1]
    assert dashboard.metadata["content_kind"] == "visual_only"
    assert dashboard.metadata["visual_region_count"] == 1
    assert dashboard.metadata["visual_regions"][0]["visual_kind"] == "image"
    assert "Dashboard Snapshot" in dashboard.content
    mock_call_vision.assert_called_once()


@patch(
    "ingestion.processors.xlsx.get_stage_model_config",
    return_value={
        "model": "gpt-5-mini",
        "max_tokens": 800,
        "temperature": None,
    },
)
@patch(
    "ingestion.processors.xlsx.tiktoken.encoding_for_model",
    return_value=_make_fake_encoding(),
)
@patch(
    "ingestion.processors.xlsx.get_xlsx_sheet_token_limit", return_value=12000
)
@patch(
    "ingestion.processors.xlsx.load_prompt",
    side_effect=_make_prompt_loader,
)
def test_process_xlsx_grid_with_chart_includes_visual_metadata(
    _prompt, _token_limit, _encoding, _model_config, tmp_path
):
    """Grid sheets with charts include both grid and visual metadata."""
    file_path = tmp_path / "mixed.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    sheet.append(["Month", "Revenue"])
    sheet.append(["Jan", 100])
    sheet.append(["Feb", 200])
    chart = BarChart()
    chart.title = "Revenue Trend"
    data_range = Reference(sheet, min_col=2, min_row=1, max_row=3)
    chart.add_data(data_range, titles_from_data=True)
    sheet.add_chart(chart, "D1")
    workbook.save(file_path)
    workbook.close()

    llm = MagicMock()
    llm.call.side_effect = [
        _make_visual_response(),
        _make_llm_response("page_like", 0.9, "Grid with chart."),
    ]

    result = process_xlsx(str(file_path), llm)

    page = result.pages[0]
    assert page.metadata["content_kind"] == "grid"
    assert page.metadata["sheet_layout_kind"] == "mixed"
    assert page.metadata["chart_count"] == 1
    assert page.metadata["visual_region_count"] == 1
    assert page.metadata["visual_regions"][0]["visual_kind"] == "chart"
    assert "## Visual Elements" in page.content
    assert "## Visual Region Descriptions" in page.content
    assert "Revenue" in page.content
    assert llm.call.call_count == 2


def test_extract_chart_title_reads_rich_text_runs():
    """Chart titles are recovered from rich text run elements."""
    chart = SimpleNamespace(
        title=SimpleNamespace(
            tx=SimpleNamespace(
                rich=SimpleNamespace(
                    p=[
                        SimpleNamespace(
                            r=[
                                SimpleNamespace(t="Q1"),
                                SimpleNamespace(t="Revenue"),
                            ],
                            fld=[],
                        )
                    ]
                )
            )
        )
    )

    assert _extract_chart_title(chart) == "Q1 Revenue"


def test_serialize_sheet_empty_chartsheet_is_empty():
    """Chartsheets with no visuals are treated as empty tabs."""
    workbook = Workbook()
    chartsheet = workbook.create_chartsheet(title="EmptyChart")

    content, metadata = _serialize_sheet(chartsheet, {})

    assert content == "## Empty Sheet"
    assert metadata["content_kind"] == "empty"
    workbook.close()


def test_build_cached_values_collects_non_none_cells():
    """Cached values map includes all non-None cell values."""
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Header"
    sheet["B1"] = 42

    values = _build_cached_values(sheet)

    assert values == {(1, 1): "Header", (1, 2): 42}
    workbook.close()


def test_serialize_sheet_prefers_cached_values_over_formulas():
    """Cached display values replace formula text in serialized content."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Calc"
    sheet["A1"] = "Revenue"
    sheet["B1"] = 100
    sheet["A2"] = "Expenses"
    sheet["B2"] = 60
    sheet["A3"] = "Margin"
    sheet["B3"] = "=B1-B2"

    cached_values = {(3, 2): 40}
    content, metadata = _serialize_sheet(sheet, cached_values)

    assert "40" in content
    assert "=B1-B2" not in content
    assert metadata["formula_cells"] == 1
    workbook.close()


def test_serialize_sheet_falls_back_to_formula_without_cache():
    """Formula text is used when no cached display value exists."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Calc"
    sheet["A1"] = "Revenue"
    sheet["B1"] = 100
    sheet["A2"] = "Total"
    sheet["B2"] = "=SUM(B1)"

    content, metadata = _serialize_sheet(sheet, {})

    assert "=SUM(B1)" in content
    assert metadata["formula_cells"] == 1
    workbook.close()


@patch(
    "ingestion.processors.xlsx.get_stage_model_config",
    return_value={
        "model": "gpt-5-mini",
        "max_tokens": 800,
        "temperature": None,
    },
)
@patch(
    "ingestion.processors.xlsx.tiktoken.encoding_for_model",
    return_value=_make_fake_encoding(),
)
@patch(
    "ingestion.processors.xlsx.get_xlsx_sheet_token_limit", return_value=12000
)
@patch("ingestion.processors.xlsx.load_prompt", return_value=_make_prompt())
def test_process_xlsx_empty_sheet_skips_llm(
    _prompt, _token_limit, _encoding, _model_config, tmp_path
):
    """Empty worksheets do not call the classifier."""
    file_path = tmp_path / "empty.xlsx"
    _save_workbook(file_path, [("Empty", [])])
    llm = MagicMock()

    result = process_xlsx(str(file_path), llm)

    assert result.total_pages == 1
    page = result.pages[0]
    assert page.content == "## Empty Sheet"
    assert page.metadata["classification"] == "empty_sheet"
    assert page.metadata["contains_dense_table"] is False
    assert page.metadata["region_count"] == 0
    assert page.metadata["dense_table_candidate_detected"] is False
    llm.call.assert_not_called()


@patch("ingestion.processors.xlsx.time.sleep")
@patch(
    "ingestion.processors.xlsx.get_xlsx_classification_retry_delay",
    return_value=2.0,
)
@patch(
    "ingestion.processors.xlsx.get_xlsx_classification_max_retries",
    return_value=3,
)
@patch(
    "ingestion.processors.xlsx.get_stage_model_config",
    return_value={
        "model": "gpt-5-mini",
        "max_tokens": 800,
        "temperature": None,
    },
)
@patch(
    "ingestion.processors.xlsx.tiktoken.encoding_for_model",
    return_value=_make_fake_encoding(),
)
@patch(
    "ingestion.processors.xlsx.get_xlsx_sheet_token_limit", return_value=12000
)
@patch("ingestion.processors.xlsx.load_prompt", return_value=_make_prompt())
def test_process_xlsx_retries_classification(
    _prompt,
    _token_limit,
    _encoding,
    _model_config,
    _retries,
    _retry_delay,
    mock_sleep,
    tmp_path,
):
    """Transient classification failures are retried."""
    file_path = tmp_path / "retry.xlsx"
    _save_workbook(
        file_path,
        [("Summary", [["Metric", "Value"], ["Revenue", 120]])],
    )
    llm = MagicMock()
    llm.call.side_effect = [
        RuntimeError("LLM unavailable"),
        _make_llm_response("page_like", 0.91, "Small printable summary."),
    ]

    result = process_xlsx(str(file_path), llm)

    assert result.total_pages == 1
    assert result.pages_succeeded == 1
    assert result.pages[0].metadata["classification"] == "page_like"
    assert llm.call.call_count == 2
    mock_sleep.assert_called_once_with(2.0)


@patch("ingestion.processors.xlsx.time.sleep")
@patch(
    "ingestion.processors.xlsx.get_xlsx_classification_retry_delay",
    return_value=1.5,
)
@patch(
    "ingestion.processors.xlsx.get_xlsx_classification_max_retries",
    return_value=2,
)
@patch(
    "ingestion.processors.xlsx.get_stage_model_config",
    return_value={
        "model": "gpt-5-mini",
        "max_tokens": 800,
        "temperature": None,
    },
)
@patch(
    "ingestion.processors.xlsx.tiktoken.encoding_for_model",
    return_value=_make_fake_encoding(),
)
@patch(
    "ingestion.processors.xlsx.get_xlsx_sheet_token_limit", return_value=12000
)
@patch("ingestion.processors.xlsx.load_prompt", return_value=_make_prompt())
def test_process_xlsx_classification_failure_aborts_workbook(
    _prompt,
    _token_limit,
    _encoding,
    _model_config,
    _retries,
    _retry_delay,
    mock_sleep,
    tmp_path,
):
    """Exhausted retries fail the workbook so it can be retried later."""
    file_path = tmp_path / "mixed.xlsx"
    _save_workbook(
        file_path,
        [
            ("Summary", [["Metric", "Value"], ["Revenue", 120]]),
            ("Broken", [["Account", "Value"], ["A", 1]]),
        ],
    )
    llm = MagicMock()
    llm.call.side_effect = [
        _make_llm_response("page_like", 0.91, "Small printable summary."),
        RuntimeError("LLM unavailable"),
        RuntimeError("LLM unavailable"),
    ]

    with pytest.raises(
        RuntimeError,
        match="sheet 'Broken': LLM unavailable",
    ):
        process_xlsx(str(file_path), llm)
    assert llm.call.call_count == 3
    mock_sleep.assert_called_once_with(1.5)


@patch(
    "ingestion.processors.xlsx.get_xlsx_classification_retry_delay",
    return_value=2.0,
)
@patch(
    "ingestion.processors.xlsx.get_xlsx_classification_max_retries",
    return_value=0,
)
def test_classify_sheet_with_retry_zero_retries_raises_runtime_error(
    _retries, _retry_delay
):
    """A zero-retry configuration fails with a clear runtime error."""
    with pytest.raises(RuntimeError, match="without a response"):
        _classify_sheet_with_retry(
            llm=MagicMock(),
            prompt=_make_prompt(),
            sheet_content="# Sheet: Summary",
            sheet_metadata={
                "sheet_name": "Summary",
                "sheet_kind": "worksheet",
                "used_range": "A1:B2",
                "row_count": 2,
                "column_count": 2,
                "row_span": 2,
                "column_span": 2,
                "blank_rows_omitted": 0,
                "blank_columns_omitted": 0,
                "non_empty_cells": 4,
                "formula_cells": 0,
                "merged_range_count": 0,
                "chart_count": 0,
                "image_count": 0,
                "max_populated_cells_in_row": 2,
                "estimated_tokens": 10,
                "token_limit": 12000,
            },
        )


@patch(
    "ingestion.processors.xlsx.get_stage_model_config",
    return_value={
        "model": "gpt-5-mini",
        "max_tokens": 800,
        "temperature": None,
    },
)
@patch(
    "ingestion.processors.xlsx.get_xlsx_sheet_token_limit", return_value=12000
)
@patch("ingestion.processors.xlsx.load_prompt", return_value=_make_prompt())
@patch(
    "ingestion.processors.xlsx.load_workbook", side_effect=OSError("bad zip")
)
def test_process_xlsx_open_failure(
    _load_workbook, _prompt, _token_limit, _model_config
):
    """Workbook open failures are surfaced as RuntimeError."""
    with pytest.raises(
        RuntimeError, match="Failed to open XLSX 'broken.xlsx'"
    ):
        process_xlsx("/tmp/broken.xlsx", MagicMock())


@patch(
    "ingestion.processors.xlsx.load_workbook",
)
def test_open_workbooks_cached_values_failure_closes_formula_workbook(
    mock_load_workbook,
):
    """Cached-value open failures close the already-open formula workbook."""
    formula_workbook = MagicMock()
    mock_load_workbook.side_effect = [
        formula_workbook,
        OSError("cached values unavailable"),
    ]

    with pytest.raises(
        RuntimeError, match="Failed to open XLSX 'broken.xlsx'"
    ):
        _open_workbooks("/tmp/broken.xlsx")
    formula_workbook.close.assert_called_once()
