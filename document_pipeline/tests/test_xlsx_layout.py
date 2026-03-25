"""Tests for XLSX layout and region detection."""

from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.worksheet.table import Table

from ingestion.processors.xlsx.layout import (
    SheetCell,
    SheetRegion,
    _estimate_long_text_ratio,
    _normalize_cell_value,
    _score_rows_with_repeated_shape,
    _split_cells_by_blank_columns,
    _split_cells_by_blank_rows,
    build_region_used_range,
    build_sheet_regions,
    collect_sheet_cells,
    detect_sheet_regions,
    score_dense_table_region,
    select_dense_table_region,
)


def test_collect_sheet_cells_preserves_original_coordinates():
    """Collected cells keep row/column coordinates and formula flags."""
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Title"
    sheet["C3"] = "=A2+B2"

    cells = collect_sheet_cells(sheet, {(3, 3): 42})
    cell_map = {(cell.row, cell.column): cell for cell in cells}

    assert cell_map[(1, 1)].value == "Title"
    assert cell_map[(3, 3)].value == "42"
    assert cell_map[(3, 3)].is_formula is True
    workbook.close()


def test_normalize_cell_value_formats_dates_and_datetimes():
    """Date-like values normalize to stable ISO-like strings."""
    assert _normalize_cell_value(False) == "FALSE"
    assert _normalize_cell_value(datetime(2026, 3, 18, 0, 0, 0)) == (
        "2026-03-18"
    )
    assert _normalize_cell_value(datetime(2026, 3, 18, 12, 30, 0)) == (
        "2026-03-18 12:30:00"
    )
    assert _normalize_cell_value(date(2026, 3, 18)) == "2026-03-18"


def test_split_cell_helpers_return_none_for_empty_input():
    """Blank-band split helpers return None for empty input."""
    assert _split_cells_by_blank_rows([]) is None
    assert _split_cells_by_blank_columns([]) is None


def test_detect_sheet_regions_splits_notes_above_table():
    """Blank row bands split top notes from the table region."""
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Portfolio Summary"
    sheet["B1"] = "Prepared by Finance"
    sheet["A3"] = "Account"
    sheet["B3"] = "Region"
    sheet["C3"] = "Amount"
    sheet.append(["A-1", "East", 100])
    sheet.append(["A-2", "West", 200])

    regions = detect_sheet_regions(collect_sheet_cells(sheet, {}))

    assert len(regions) == 2
    assert build_region_used_range(regions[0]) == "A1:B1"
    assert build_region_used_range(regions[1]) == "A3:C5"
    workbook.close()


def test_detect_sheet_regions_returns_empty_for_no_cells():
    """Empty sheets produce no detected regions."""
    assert not detect_sheet_regions([])


def test_detect_sheet_regions_splits_side_notes_from_table():
    """Blank column bands split side notes from the table region."""
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Owner"
    sheet["B1"] = "Treasury"
    sheet["A2"] = "As of"
    sheet["B2"] = "2026-03-18"
    sheet["D1"] = "Account"
    sheet["E1"] = "Region"
    sheet["F1"] = "Amount"
    sheet["D2"] = "A-1"
    sheet["E2"] = "East"
    sheet["F2"] = 100
    sheet["D3"] = "A-2"
    sheet["E3"] = "West"
    sheet["F3"] = 200

    regions = detect_sheet_regions(collect_sheet_cells(sheet, {}))

    assert len(regions) == 2
    assert build_region_used_range(regions[0]) == "A1:B2"
    assert build_region_used_range(regions[1]) == "D1:F3"
    workbook.close()


def test_select_dense_table_region_prefers_repeated_row_shape():
    """Selection prefers the larger repeated-row table region."""
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Scope"
    sheet["B1"] = "Quarterly commentary for internal review only"
    sheet["A3"] = "Account"
    sheet["B3"] = "Region"
    sheet["C3"] = "Quarter"
    sheet["D3"] = "Amount"
    for row in (
        ["A-1", "East", "Q1", 100],
        ["A-2", "West", "Q1", 200],
        ["A-3", "East", "Q2", 300],
        ["A-4", "West", "Q2", 400],
    ):
        sheet.append(row)

    regions = detect_sheet_regions(collect_sheet_cells(sheet, {}))
    selected = select_dense_table_region(regions)

    assert selected is not None
    assert build_region_used_range(selected) == "A3:D7"
    assert selected.region_type == "dense_table_candidate"
    assert selected.dense_score > score_dense_table_region(regions[0])
    workbook.close()


def test_score_helpers_handle_empty_region_rows():
    """Empty row payloads produce safe zero or full penalties."""
    region = SheetRegion(
        region_id="region_1",
        region_type="framing",
        min_row=1,
        min_col=1,
        max_row=2,
        max_col=2,
        non_empty_cells=0,
        row_count=2,
        column_count=2,
        row_numbers=[1, 2],
        column_numbers=[1, 2],
        rows=[],
    )

    assert _score_rows_with_repeated_shape(region) == 0.0
    assert _estimate_long_text_ratio(region) == 1.0


def test_detect_sheet_regions_handles_two_tables_on_one_sheet():
    """Blank row bands split multiple tables on the same sheet."""
    workbook = Workbook()
    sheet = workbook.active
    for row in (
        ["Account", "Amount"],
        ["A-1", 100],
        ["A-2", 200],
        [None, None],
        ["Region", "Balance"],
        ["East", 300],
        ["West", 400],
    ):
        sheet.append(row)

    regions = detect_sheet_regions(collect_sheet_cells(sheet, {}))

    assert len(regions) == 2
    assert [build_region_used_range(region) for region in regions] == [
        "A1:B3",
        "A5:B7",
    ]
    workbook.close()


def test_detect_sheet_regions_prefers_native_excel_table_bounds():
    """Native Excel tables win over heuristic dense regions."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Account", "Region", "Amount"])
    sheet.append(["A-1", None, 100])
    sheet.append(["A-2", "West", 200])
    sheet.append(["A-3", "East", 300])
    native_table = Table(displayName="ExposureTable", ref="A1:C4")
    sheet.add_table(native_table)
    sheet["E1"] = "Notes"
    sheet["F1"] = "Quarter-end snapshot"

    regions = build_sheet_regions(sheet, {})
    selected = select_dense_table_region(regions)

    assert len(regions) == 2
    assert selected is not None
    assert selected.region_type == "native_table"
    assert selected.native_table_name == "ExposureTable"
    assert build_region_used_range(selected) == "A1:C4"
    workbook.close()


def test_select_dense_table_region_prefers_larger_dense_region_over_notes():
    """Tiny native note tables do not outrank a materially denser export."""
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = "Key"
    sheet["B1"] = "Value"
    sheet["A2"] = "Owner"
    sheet["B2"] = "Treasury"
    sheet.add_table(Table(displayName="NotesTable", ref="A1:B2"))
    sheet["A5"] = "Account"
    sheet["B5"] = "Region"
    sheet["C5"] = "Amount"
    sheet["A6"] = "A-1"
    sheet["B6"] = "East"
    sheet["C6"] = 100
    sheet["A7"] = "A-2"
    sheet["B7"] = "West"
    sheet["C7"] = 200
    sheet["A8"] = "A-3"
    sheet["B8"] = "North"
    sheet["C8"] = 300
    sheet["A9"] = "A-4"
    sheet["B9"] = "South"
    sheet["C9"] = 400

    regions = build_sheet_regions(sheet, {})
    selected = select_dense_table_region(regions)

    assert selected is not None
    assert selected.region_type == "dense_table_candidate"
    assert build_region_used_range(selected) == "A5:C9"
    workbook.close()


def test_select_dense_table_region_returns_none_for_low_score():
    """Selection returns None when no region looks dense enough."""
    region = SheetRegion(
        region_id="region_1",
        region_type="framing",
        min_row=1,
        min_col=1,
        max_row=1,
        max_col=1,
        non_empty_cells=1,
        row_count=1,
        column_count=1,
        row_numbers=[1],
        column_numbers=[1],
        rows=[{1: "Note"}],
    )

    assert select_dense_table_region([region]) is None


def test_select_dense_table_region_returns_none_for_empty_input():
    """Selection returns None for an empty region list."""
    assert select_dense_table_region([]) is None


def test_build_region_used_range_single_cell():
    """Single-cell regions stay in simple A1 notation."""
    region = SheetRegion(
        region_id="region_1",
        region_type="framing",
        min_row=2,
        min_col=3,
        max_row=2,
        max_col=3,
        non_empty_cells=1,
        row_count=1,
        column_count=1,
        row_numbers=[2],
        column_numbers=[3],
        rows=[{3: "Value"}],
    )

    assert build_region_used_range(region) == "C2"


def test_split_cell_helpers_keep_single_band_unsplit():
    """Single populated bands are not split further."""
    cells = [
        SheetCell(row=1, column=1, value="A", is_formula=False),
        SheetCell(row=1, column=2, value="B", is_formula=False),
    ]

    assert _split_cells_by_blank_rows(cells) is None
    assert _split_cells_by_blank_columns(cells) is None
