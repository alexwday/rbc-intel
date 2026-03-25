"""Programmatic column analysis for dense table content."""

import re
from collections import Counter
from typing import Any, List

from openpyxl.utils import get_column_letter

from .types import ColumnProfile, TableEDA

_DATE_PATTERN = re.compile(r"^\d{4}-\d{2}-\d{2}(?:\s\d{2}:\d{2}:\d{2})?$")
_NUMERIC_STRIP = re.compile(r"[$,%]")
_PAREN_NEGATIVE = re.compile(r"^\(([0-9,.]+)\)$")
_PIPE_ROW = re.compile(r"^\|.+\|$")
_SEPARATOR_CELL = re.compile(r"^:?-+:?$")
_USED_RANGE_LINE = re.compile(r"^- Used range: (.+)$", re.MULTILINE)
_SUPPORTED_FORMATS = frozenset({"xlsx_markdown"})


def _estimate_tokens(text: str) -> int:
    """Estimate token count via word-count heuristic. Returns: int."""
    return max(1, int(len(text.split()) * 1.3))


def _unescape_cell(value: str) -> str:
    """Reverse pipe-table escaping. Params: value. Returns: str."""
    return value.replace("\\|", "|").replace("<br>", "\n")


def _escape_sample_cell(value: str) -> str:
    """Escape a value for sample row markdown. Params: value. Returns: str."""
    return value.replace("|", "\\|").replace("\n", "<br>")


def _split_pipe_row(line: str) -> list[str]:
    """Split a pipe table row into trimmed cell values."""
    stripped = line.strip()
    if stripped.startswith("|"):
        stripped = stripped[1:]
    if stripped.endswith("|"):
        stripped = stripped[:-1]
    parts = re.split(r"(?<!\\)\|", stripped)
    return [p.strip() for p in parts]


def _is_separator_row(cells: list[str]) -> bool:
    """Check if parsed cells form a --- separator row."""
    non_empty = [c for c in cells if c]
    if not non_empty:
        return False
    return all(_SEPARATOR_CELL.match(c) for c in non_empty)


def _is_pipe_row(line: str) -> bool:
    """Check if a line is a pipe table row. Returns: bool."""
    return bool(_PIPE_ROW.match(line.strip()))


def _parse_numeric(value: str) -> float | None:
    """Try to parse a value as a number. Returns: float | None."""
    cleaned = value.strip()
    if not cleaned:
        return None
    match = _PAREN_NEGATIVE.match(cleaned)
    if match:
        cleaned = "-" + match.group(1)
    cleaned = _NUMERIC_STRIP.sub("", cleaned)
    if not cleaned or cleaned == "-":
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _detect_type(value: str) -> str:
    """Detect cell data type. Params: value. Returns: str."""
    stripped = value.strip()
    if not stripped:
        return "null"
    if stripped.upper() in ("TRUE", "FALSE"):
        return "boolean"
    if _DATE_PATTERN.match(stripped):
        return "date"
    if _parse_numeric(stripped) is not None:
        return "numeric"
    return "text"


def _classify_column_dtype(type_counts: dict[str, int]) -> str:
    """Determine column dtype from cell type counts. Returns: str."""
    non_null = {k: v for k, v in type_counts.items() if k != "null"}
    if not non_null:
        return "text"
    if len(non_null) == 1:
        return next(iter(non_null))
    return "mixed"


def _compute_numeric_stats(
    values: List[float],
) -> dict[str, Any]:
    """Compute min/max/mean for numeric values. Returns: dict."""
    if not values:
        return {}
    return {
        "min": min(values),
        "max": max(values),
        "mean": round(sum(values) / len(values), 4),
    }


def _compute_date_stats(values: List[str]) -> dict[str, Any]:
    """Compute min/max for date strings. Returns: dict."""
    if not values:
        return {}
    sorted_dates = sorted(values)
    return {"min": sorted_dates[0], "max": sorted_dates[-1]}


def _compute_text_stats(
    values: List[str], unique_count: int
) -> dict[str, Any]:
    """Compute length stats and optional distribution. Returns: dict."""
    if not values:
        return {}
    lengths = [len(v) for v in values]
    stats: dict[str, Any] = {
        "min_length": min(lengths),
        "max_length": max(lengths),
        "avg_length": round(sum(lengths) / len(lengths), 1),
    }
    if unique_count <= 20:
        distribution: dict[str, int] = {}
        for v in values:
            distribution[v] = distribution.get(v, 0) + 1
        stats["value_distribution"] = distribution
    return stats


def _compute_boolean_stats(
    values: List[str],
) -> dict[str, Any]:
    """Compute true/false counts. Returns: dict."""
    true_count = sum(1 for v in values if v.upper() == "TRUE")
    return {
        "true_count": true_count,
        "false_count": len(values) - true_count,
    }


def _pick_samples(values: List[str], count: int = 5) -> List[str]:
    """Pick up to count distinct non-empty values. Returns: list."""
    seen: set[str] = set()
    samples: list[str] = []
    for v in values:
        if v.strip() and v not in seen:
            seen.add(v)
            samples.append(v)
            if len(samples) >= count:
                break
    return samples


def _profile_column(
    position: str, name: str, values: List[str]
) -> ColumnProfile:
    """Build a ColumnProfile from raw cell values.

    Params:
        position: Column letter (e.g. "A")
        name: Column header text
        values: Cell values for this column (excluding header)

    Returns:
        ColumnProfile with detected type, stats, and samples

    Example:
        >>> col = _profile_column("A", "Amount", ["100", "200"])
        >>> col.dtype
        'numeric'
    """
    non_null = [v for v in values if v.strip()]
    null_count = len(values) - len(non_null)
    unique_count = len(set(non_null))

    type_counts: dict[str, int] = {}
    for v in values:
        dtype = _detect_type(v)
        type_counts[dtype] = type_counts.get(dtype, 0) + 1

    column_dtype = _classify_column_dtype(type_counts)

    if column_dtype == "numeric":
        parsed: list[float] = []
        for v in non_null:
            num = _parse_numeric(v)
            if num is not None:
                parsed.append(num)
        stats = _compute_numeric_stats(parsed)
    elif column_dtype == "date":
        date_vals = [
            v.strip() for v in non_null if _DATE_PATTERN.match(v.strip())
        ]
        stats = _compute_date_stats(date_vals)
    elif column_dtype == "boolean":
        stats = _compute_boolean_stats(non_null)
    else:
        stats = _compute_text_stats(non_null, unique_count)

    return ColumnProfile(
        name=name,
        position=position,
        dtype=column_dtype,
        stats=stats,
        sample_values=_pick_samples(non_null),
        non_null_count=len(non_null),
        null_count=null_count,
        unique_count=unique_count,
    )


def _extract_sample_rows(
    data_lines: List[str], head: int = 5, tail: int = 3
) -> List[str]:
    """Get representative rows (first head + last tail). Returns: list."""
    if len(data_lines) <= head + tail:
        return list(data_lines)
    return data_lines[:head] + data_lines[-tail:]


def _extract_used_range(content: str) -> str:
    """Extract the worksheet used range from serialized content."""
    match = _USED_RANGE_LINE.search(content)
    if match is None:
        return ""
    return match.group(1).strip()


def _value_shape(value: str) -> str:
    """Collapse a value into a simple shape signature. Returns: str."""
    parts: list[str] = []
    for char in value.strip():
        if char.isalpha():
            token = "A"
        elif char.isdigit():
            token = "0"
        elif char.isspace():
            token = " "
        else:
            token = char
        if not parts or parts[-1] != token:
            parts.append(token)
    return "".join(parts)


def _dominant_non_null_type(values: List[str]) -> str:
    """Return the dominant non-null detected type for a value list."""
    counts = Counter(
        _detect_type(value)
        for value in values
        if _detect_type(value) != "null"
    )
    if not counts:
        return "null"
    return counts.most_common(1)[0][0]


def _is_label_like(value: str) -> bool:
    """Heuristically detect a header-style text label."""
    stripped = value.strip()
    if not stripped:
        return False
    return (
        " " in stripped
        or "_" in stripped
        or stripped.isupper()
        or stripped.endswith(":")
    )


def _find_table_bounds(
    lines: List[str],
) -> tuple[int, int]:
    """Find first and last pipe table row indices. Returns: (start, end)."""
    start = -1
    end = -1
    for i, line in enumerate(lines):
        if _is_pipe_row(line):
            if start == -1:
                start = i
            end = i
    return start, end


def _build_framing_context(
    lines: List[str], table_start: int, table_end: int
) -> str:
    """Build framing context from lines outside the pipe table."""
    before = "\n".join(lines[:table_start]).strip()
    after = "\n".join(lines[table_end + 1 :]).strip()
    parts = [p for p in (before, after) if p]
    return "\n\n".join(parts)


def _collect_data_lines(
    table_lines: List[str],
) -> List[str]:
    """Extract data rows from pipe table (skip separator)."""
    data: list[str] = []
    for tline in table_lines[1:]:
        cells = _split_pipe_row(tline)
        if not _is_separator_row(cells):
            data.append(tline)
    return data


def _parse_column_values(
    value_rows: List[str],
    positions: List[str],
    offset: int,
) -> list[list[str]]:
    """Parse data rows into per-column value lists."""
    col_values: list[list[str]] = [[] for _ in positions]
    for row_line in value_rows:
        cells = _split_pipe_row(row_line)
        row_data = [_unescape_cell(c) for c in cells[offset:]]
        for col_idx in range(len(positions)):
            if col_idx < len(row_data):
                col_values[col_idx].append(row_data[col_idx])
            else:
                col_values[col_idx].append("")
    return col_values


def _empty_eda(framing: str, content: str) -> TableEDA:
    """Build a TableEDA with no columns or data rows."""
    return TableEDA(
        row_count=0,
        columns=[],
        header_row=0,
        framing_context=framing,
        sample_rows=[],
        token_count=_estimate_tokens(content),
    )


def _empty_region_eda(
    framing_context: str,
    token_source: str,
    used_range: str,
    source_region_id: str,
) -> TableEDA:
    """Build an empty region-based TableEDA result."""
    return TableEDA(
        row_count=0,
        columns=[],
        header_row=0,
        framing_context=framing_context,
        sample_rows=[],
        token_count=_estimate_tokens(token_source),
        used_range=used_range,
        header_mode="headerless",
        source_region_id=source_region_id,
    )


def _parse_table_structure(
    table_lines: List[str],
    data_lines: List[str],
) -> tuple[List[str], int, int, List[str]]:
    """Parse column positions, offset, header row, and names."""
    header_cells = _split_pipe_row(table_lines[0])
    has_row_col = len(header_cells) > 0 and header_cells[0].strip() == "Row"
    offset = 1 if has_row_col else 0
    positions = header_cells[offset:]
    first_cells = _split_pipe_row(data_lines[0])
    header_row_num = _parse_header_row_num(first_cells, has_row_col)
    col_names = _pad_col_names(first_cells, offset, len(positions))
    return positions, offset, header_row_num, col_names


def _score_header_cell(
    first_value: str,
    later_values: list[str],
) -> tuple[int, int]:
    """Score one column's first-row value as header-like or data-like."""
    if not first_value or not later_values:
        return 0, 0

    data_score = 0
    header_score = 0
    first_type = _detect_type(first_value)
    dominant_type = _dominant_non_null_type(later_values)
    if first_value in later_values:
        data_score += 2

    if dominant_type == first_type:
        if first_type == "text":
            later_shapes = {_value_shape(value) for value in later_values[:5]}
            first_shape = _value_shape(first_value)
            if first_shape in later_shapes:
                data_score += 1
            if _is_label_like(first_value):
                header_score += 1
            if first_value not in later_values:
                unique_later = set(later_values[:5])
                if len(unique_later) < len(later_values[:5]):
                    header_score += 1
        else:
            data_score += 1
        return data_score, header_score

    if first_type == "text" and dominant_type in (
        "numeric",
        "date",
        "boolean",
    ):
        header_score += 2
    else:
        header_score += 1
    return data_score, header_score


def detect_header_mode(
    rows: list[dict[int, str]],
    column_numbers: list[int],
) -> str:
    """Detect whether the first region row is a header or data row."""
    if not rows:
        return "header_row"

    first_values = [
        rows[0].get(column_number, "") for column_number in column_numbers
    ]
    if len(rows) == 1:
        if _looks_like_data_row(first_values):
            return "headerless"
        return "header_row"

    if _looks_like_data_row(first_values):
        return "headerless"

    header_score = 0
    data_score = 0
    non_empty = [value for value in first_values if value.strip()]
    if non_empty and len(set(non_empty)) == len(non_empty):
        header_score += 1

    for column_number in column_numbers:
        first_value = rows[0].get(column_number, "").strip()
        later_values = [
            row.get(column_number, "").strip()
            for row in rows[1:]
            if row.get(column_number, "").strip()
        ]
        data_delta, header_delta = _score_header_cell(
            first_value=first_value,
            later_values=later_values,
        )
        data_score += data_delta
        header_score += header_delta

    if data_score >= header_score + 1:
        return "headerless"
    return "header_row"


def build_column_names(
    rows: list[dict[int, str]],
    column_numbers: list[int],
    header_mode: str,
    row_numbers: list[int] | None = None,
) -> tuple[int, list[str], list[dict[int, str]]]:
    """Build column names and data rows from a region row set."""
    positions = [
        get_column_letter(column_number) for column_number in column_numbers
    ]
    if not rows:
        return 0, positions, []

    if header_mode == "headerless":
        return 0, positions, rows

    header_row_num = 0
    if row_numbers:
        header_row_num = row_numbers[0]

    header_values = [
        rows[0].get(column_number, "") for column_number in column_numbers
    ]
    names = [_unescape_cell(value) for value in header_values]
    return header_row_num, names, rows[1:]


def _coerce_column_numbers(region: dict[str, Any]) -> list[int]:
    """Extract sorted column numbers from serialized region metadata."""
    raw_numbers = region.get("column_numbers", [])
    numbers = [int(value) for value in raw_numbers if isinstance(value, int)]
    if numbers:
        return sorted(numbers)

    numbers = []
    rows = region.get("rows", [])
    if isinstance(rows, list):
        for row in rows:
            if not isinstance(row, dict):
                continue
            cells = row.get("cells", [])
            if not isinstance(cells, list):
                continue
            for cell in cells:
                if not isinstance(cell, dict):
                    continue
                column_number = cell.get("column_number")
                if isinstance(column_number, int):
                    numbers.append(column_number)
    return sorted(set(numbers))


def _coerce_region_rows(
    region: dict[str, Any],
) -> tuple[list[int], list[dict[int, str]]]:
    """Coerce serialized region row metadata into row maps."""
    explicit_row_numbers = [
        int(value)
        for value in region.get("row_numbers", [])
        if isinstance(value, int)
    ]
    rows_by_number: dict[int, dict[int, str]] = {}
    raw_rows = region.get("rows", [])
    if isinstance(raw_rows, list):
        for row in raw_rows:
            if not isinstance(row, dict):
                continue
            row_number = row.get("row_number")
            if not isinstance(row_number, int):
                continue
            cells = row.get("cells", [])
            if not isinstance(cells, list):
                continue
            row_map: dict[int, str] = {}
            for cell in cells:
                if not isinstance(cell, dict):
                    continue
                column_number = cell.get("column_number")
                if not isinstance(column_number, int):
                    continue
                value = cell.get("value", "")
                row_map[column_number] = "" if value is None else str(value)
            rows_by_number[row_number] = row_map

    ordered_numbers = explicit_row_numbers or sorted(rows_by_number)
    for row_number in rows_by_number:
        if row_number not in ordered_numbers:
            ordered_numbers.append(row_number)
    return ordered_numbers, [
        rows_by_number.get(row_number, {}) for row_number in ordered_numbers
    ]


def _parse_region_column_values(
    value_rows: list[dict[int, str]],
    column_numbers: list[int],
) -> list[list[str]]:
    """Parse region rows into per-column value lists."""
    col_values: list[list[str]] = [[] for _ in column_numbers]
    for row in value_rows:
        for index, column_number in enumerate(column_numbers):
            col_values[index].append(row.get(column_number, ""))
    return col_values


def _build_region_sample_rows(
    row_numbers: list[int],
    rows: list[dict[int, str]],
    column_numbers: list[int],
) -> list[str]:
    """Build sample pipe rows from region row dictionaries."""
    indexed_rows = list(zip(row_numbers, rows))
    if len(indexed_rows) > 8:
        indexed_rows = indexed_rows[:5] + indexed_rows[-3:]

    sample_rows: list[str] = []
    for row_number, row in indexed_rows:
        rendered = [
            _escape_sample_cell(row.get(column_number, ""))
            for column_number in column_numbers
        ]
        sample_rows.append(f"| {row_number} | {' | '.join(rendered)} |")
    return sample_rows


def _build_profile_columns(
    positions: list[str],
    col_names: list[str],
    col_values: list[list[str]],
) -> list[ColumnProfile]:
    """Build profiled columns from aligned name/value lists."""
    return [
        _profile_column(
            positions[index],
            col_names[index] if index < len(col_names) else positions[index],
            col_values[index] if index < len(col_values) else [],
        )
        for index in range(len(positions))
    ]


def _build_region_eda(
    region: dict[str, Any],
    framing_context: str,
    token_source: str,
) -> TableEDA:
    """Build a TableEDA from serialized region metadata."""
    used_range = str(region.get("used_range", "")).strip()
    source_region_id = str(region.get("region_id", "")).strip()
    column_numbers = _coerce_column_numbers(region)
    row_numbers, rows = _coerce_region_rows(region)
    token_text = (
        token_source or framing_context or used_range or source_region_id
    )
    if not column_numbers or not rows:
        return _empty_region_eda(
            framing_context=framing_context,
            token_source=token_text,
            used_range=used_range,
            source_region_id=source_region_id,
        )

    header_mode = detect_header_mode(rows, column_numbers)
    header_row, col_names, value_rows = build_column_names(
        rows=rows,
        column_numbers=column_numbers,
        header_mode=header_mode,
        row_numbers=row_numbers,
    )
    positions = [
        get_column_letter(column_number) for column_number in column_numbers
    ]
    sample_row_numbers = (
        row_numbers if header_mode == "headerless" else row_numbers[1:]
    )

    return TableEDA(
        row_count=len(value_rows),
        columns=_build_profile_columns(
            positions=positions,
            col_names=_disambiguate_names(col_names, positions),
            col_values=_parse_region_column_values(
                value_rows,
                column_numbers,
            ),
        ),
        header_row=header_row,
        framing_context=framing_context,
        sample_rows=_build_region_sample_rows(
            sample_row_numbers,
            value_rows,
            column_numbers,
        ),
        token_count=_estimate_tokens(token_text),
        used_range=used_range,
        header_mode=header_mode,
        source_region_id=source_region_id,
    )


def run_table_eda_from_region(
    region: dict[str, Any],
    framing_context: str = "",
    token_source: str = "",
) -> TableEDA:
    """Profile a dense table directly from region metadata."""
    return _build_region_eda(
        region=region,
        framing_context=framing_context,
        token_source=token_source,
    )


def _resolve_markdown_values(
    table_lines: list[str],
    data_lines: list[str],
) -> tuple[list[str], int, int, list[str], list[str], str]:
    """Resolve markdown table headers and data rows."""
    positions, offset, header_row_num, col_names = _parse_table_structure(
        table_lines,
        data_lines,
    )
    if _looks_like_data_row(col_names):
        return (
            positions,
            offset,
            0,
            list(positions),
            data_lines,
            "headerless",
        )
    return (
        positions,
        offset,
        header_row_num,
        col_names,
        data_lines[1:],
        "header_row",
    )


def run_table_eda(
    content: str, source_format: str = "xlsx_markdown"
) -> TableEDA:
    """Profile dense table columns from serialized markdown.

    Parses the markdown pipe table, separates framing context
    (metadata header, visual elements), detects column types,
    and computes per-column statistics.

    Params:
        content: Serialized markdown from extraction
        source_format: Content format ("xlsx_markdown")

    Returns:
        TableEDA with column profiles, framing context, samples

    Example:
        >>> eda = run_table_eda(sheet_content)
        >>> eda.row_count
        100
    """
    if source_format not in _SUPPORTED_FORMATS:
        raise ValueError(f"Unsupported source_format: {source_format}")

    lines = content.splitlines()
    table_start, table_end = _find_table_bounds(lines)

    if table_start == -1:
        return _empty_eda(content.strip(), content)

    framing = _build_framing_context(lines, table_start, table_end)
    table_lines = lines[table_start : table_end + 1]
    data_lines = _collect_data_lines(table_lines)
    if not data_lines:
        return _empty_eda(framing, content)

    (
        positions,
        offset,
        header_row_num,
        col_names,
        value_rows,
        header_mode,
    ) = _resolve_markdown_values(
        table_lines,
        data_lines,
    )

    col_names = _disambiguate_names(col_names, positions)
    col_values = _parse_column_values(value_rows, positions, offset)

    return TableEDA(
        row_count=len(value_rows),
        columns=_build_profile_columns(positions, col_names, col_values),
        header_row=header_row_num,
        framing_context=framing,
        sample_rows=_extract_sample_rows(value_rows),
        token_count=_estimate_tokens(content),
        used_range=_extract_used_range(content),
        header_mode=header_mode,
        source_region_id="",
    )


def _parse_header_row_num(first_cells: list[str], has_row_col: bool) -> int:
    """Extract the Excel row number from first data row."""
    if not has_row_col or not first_cells:
        return 0
    try:
        return int(first_cells[0])
    except ValueError:
        return 0


def _pad_col_names(
    first_cells: list[str], offset: int, num_positions: int
) -> list[str]:
    """Extract and pad column names to match positions."""
    names = [_unescape_cell(c) for c in first_cells[offset:]]
    while len(names) < num_positions:
        names.append("")
    return names


def _looks_like_data_row(values: List[str]) -> bool:
    """Check if values look like data rather than column headers.

    Returns True when a strict majority of non-null values are
    numeric, date, or boolean — types unlikely to be column names.

    Params: values (list[str]). Returns: bool.
    """
    non_null = [v for v in values if v.strip()]
    if not non_null:
        return False
    non_text_count = sum(1 for v in non_null if _detect_type(v) != "text")
    return non_text_count > len(non_null) / 2


def _disambiguate_names(names: List[str], positions: List[str]) -> List[str]:
    """Append column position to duplicate names for uniqueness.

    Params:
        names: Column names (may contain duplicates)
        positions: Column letters (A, B, C, ...)

    Returns:
        list[str] — names with duplicates suffixed by position

    Example:
        >>> _disambiguate_names(["ID", "Amount", "Amount"], ["A", "B", "C"])
        ['ID', 'Amount (B)', 'Amount (C)']
    """
    counts: dict[str, int] = {}
    for name in names:
        counts[name] = counts.get(name, 0) + 1

    duplicated = {name for name, count in counts.items() if count > 1}
    if not duplicated:
        return names

    result: list[str] = []
    for idx, name in enumerate(names):
        if name in duplicated:
            pos = positions[idx] if idx < len(positions) else ""
            result.append(f"{name} ({pos})" if pos else name)
        else:
            result.append(name)
    return result
