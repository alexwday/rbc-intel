"""XLSX processor — classifies each sheet for dense-table handling."""

import base64
import json
import logging
import time
from datetime import date, datetime
from importlib import import_module
from itertools import zip_longest
from pathlib import Path
from typing import Any

import openai
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries, range_to_tuple
from openpyxl.worksheet.worksheet import Worksheet

from ..connections.llm import LLMClient
from ..utils.config import (
    get_stage_model_config,
    get_xlsx_classification_max_retries,
    get_xlsx_classification_retry_delay,
    get_xlsx_sheet_token_limit,
)
from ..utils.file_types import ExtractionResult, PageResult
from ..utils.prompt_loader import load_prompt
from ..utils.xlsx_layout import (
    SheetRegion,
    build_region_used_range,
    build_sheet_regions,
    select_dense_table_region,
)

logger = logging.getLogger(__name__)

_MAX_RETRIES = 3
_RETRY_DELAY_S = 2.0

RETRYABLE_ERRORS = (
    openai.RateLimitError,
    openai.APITimeoutError,
    openai.APIConnectionError,
    openai.InternalServerError,
)


def parse_tool_arguments(response: dict) -> dict:
    """Extract parsed tool-call arguments from an LLM response.

    Params: response (dict). Returns: dict.
    """
    choices = response.get("choices")
    if not isinstance(choices, list) or not choices:
        raise ValueError("LLM response missing choices")

    message = choices[0].get("message")
    if not isinstance(message, dict):
        raise ValueError("LLM response missing message payload")

    tool_calls = message.get("tool_calls")
    if not isinstance(tool_calls, list) or not tool_calls:
        raise ValueError("LLM response missing tool calls")

    function_data = tool_calls[0].get("function")
    if not isinstance(function_data, dict):
        raise ValueError("LLM response missing function payload")

    arguments = function_data.get("arguments")
    if not isinstance(arguments, str):
        raise ValueError("LLM response missing function arguments")

    parsed = json.loads(arguments)
    if not isinstance(parsed, dict):
        raise ValueError("LLM tool arguments must decode to an object")
    return parsed


def parse_vision_response(response: dict) -> tuple[str, str]:
    """Parse the first tool call into page title and content.

    Params: response (dict). Returns: tuple[str, str].
    """
    parsed = parse_tool_arguments(response)
    page_title = parsed.get("page_title")
    content = parsed.get("content")
    if not isinstance(page_title, str) or not isinstance(content, str):
        raise ValueError(
            "LLM tool arguments missing string page_title/content"
        )
    return page_title, content


def call_vision(llm, img_bytes, prompt, tool_choice, detail="auto") -> tuple:
    """Send a page image to the LLM and extract content.

    Builds a multi-modal message with system prompt and user
    prompt containing the base64 image. Parses tool call
    response for page_title and content. Retries on transient
    OpenAI errors.

    Params:
        llm: LLMClient instance
        img_bytes: PNG image bytes
        prompt: Loaded prompt dict with system_prompt,
            user_prompt, stage, tools, tool_choice
        tool_choice: Tool choice override (prompt value
            or constructed dict)
        detail: Vision detail level — "auto" (default),
            "low", or "high"

    Returns:
        tuple of (page_title, content)

    Example:
        >>> title, md = call_vision(llm, img, p, "required")
        >>> title
        "Q3 Financial Results"
    """
    b64 = base64.b64encode(img_bytes).decode()

    messages = [
        {"role": "system", "content": prompt["system_prompt"]},
        {
            "role": "user",
            "content": [
                {
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:image/png;base64,{b64}",
                        "detail": detail,
                    },
                },
                {
                    "type": "text",
                    "text": prompt["user_prompt"],
                },
            ],
        },
    ]

    for attempt in range(1, _MAX_RETRIES + 1):
        try:
            response = llm.call(
                messages=messages,
                stage=prompt["stage"],
                tools=prompt["tools"],
                tool_choice=tool_choice,
            )
            return parse_vision_response(response)

        except RETRYABLE_ERRORS as exc:
            if attempt == _MAX_RETRIES:
                logger.error(
                    "Vision call failed after %d attempts: %s",
                    _MAX_RETRIES,
                    exc,
                )
                raise
            wait = _RETRY_DELAY_S * attempt
            logger.warning(
                "Vision retry %d/%d after %.1fs: %s",
                attempt,
                _MAX_RETRIES,
                wait,
                exc,
            )
            time.sleep(wait)
    raise RuntimeError("Vision call exited retry loop without a response")


_PREVIEW_HEAD_LINES = 24
_PREVIEW_TAIL_LINES = 10
_PREVIEW_MAX_LINE_LENGTH = 300
_REGION_PREVIEW_MAX_ROWS = 12
_REGION_PREVIEW_MAX_COLUMNS = 10
_REGION_PREVIEW_MAX_VALUE_LENGTH = 80
_VISUAL_SERIES_POINT_LIMIT = 12
_RETRYABLE_CLASSIFICATION_ERRORS = (
    openai.OpenAIError,
    RuntimeError,
    ValueError,
)


class _FallbackEncoding:
    """Minimal tokenizer fallback when tiktoken is unavailable."""

    def encode(self, text: str) -> list[int]:
        """Approximate BPE token count using ~4 chars per token."""
        return list(range(max(1, len(text) // 4)))

    def decode(self, tokens: list[int]) -> str:
        """Decode is unsupported on the fallback tokenizer."""
        raise NotImplementedError("Fallback tokenizer does not support decode")


class _FallbackTiktoken:
    """Compatibility shim exposing the tiktoken methods this module uses."""

    def encoding_for_model(self, _model: str) -> _FallbackEncoding:
        """Return a coarse tokenizer for the requested model."""
        return _FallbackEncoding()

    def get_encoding(self, _name: str) -> _FallbackEncoding:
        """Return a coarse tokenizer for the requested encoding."""
        return _FallbackEncoding()


def _load_tiktoken() -> Any:
    """Import tiktoken when available and otherwise use a local fallback."""
    try:
        return import_module("tiktoken")
    except ImportError:
        return _FallbackTiktoken()


tiktoken = _load_tiktoken()


def _build_base_sheet_metadata(
    sheet_name: str,
    visual_metadata: dict[str, Any],
) -> dict[str, Any]:
    """Build common metadata shared by all worksheet serializations."""
    return {
        "sheet_name": sheet_name,
        "sheet_kind": visual_metadata["sheet_kind"],
        "chart_count": visual_metadata["chart_count"],
        "image_count": visual_metadata["image_count"],
        "chart_titles": visual_metadata["chart_titles"],
    }


def _build_empty_sheet_metadata(
    sheet_name: str,
    visual_metadata: dict[str, Any],
) -> dict[str, Any]:
    """Build metadata for a worksheet with no populated cells."""
    metadata = _build_base_sheet_metadata(sheet_name, visual_metadata)
    metadata.update(
        {
            "content_kind": "empty",
            "used_range": "",
            "row_count": 0,
            "column_count": 0,
            "row_span": 0,
            "column_span": 0,
            "blank_rows_omitted": 0,
            "blank_columns_omitted": 0,
            "non_empty_cells": 0,
            "formula_cells": 0,
            "merged_range_count": 0,
            "max_populated_cells_in_row": 0,
        }
    )
    return metadata


def _normalize_cell_value(value: Any) -> str:
    """Normalize a worksheet cell value into stable text."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value).replace("\r\n", "\n").replace("\r", "\n")


def _escape_table_text(value: str) -> str:
    """Escape markdown table delimiters in cell text."""
    return value.replace("|", "\\|").replace("\n", "<br>")


def _extract_chart_title(chart: Any) -> str:
    """Extract plain text from an openpyxl chart title when available."""
    return _extract_title_text(getattr(chart, "title", None))


def _extract_title_text(title: Any) -> str:
    """Extract plain text from an openpyxl title-like object."""
    if title is None:
        return ""

    if isinstance(title, str):
        return title.strip()

    text_runs: list[str] = []
    text = getattr(title, "tx", None)
    rich_text = getattr(text, "rich", None)
    paragraphs = getattr(rich_text, "p", [])
    for paragraph in paragraphs:
        for run in getattr(paragraph, "r", []) or []:
            run_text = getattr(run, "t", "")
            if run_text:
                text_runs.append(str(run_text))
        for field in getattr(paragraph, "fld", []) or []:
            field_text = getattr(field, "t", "")
            if field_text:
                text_runs.append(str(field_text))
    return " ".join(part.strip() for part in text_runs if part).strip()


def _extract_anchor_position(anchor: Any) -> dict[str, Any]:
    """Extract row/column/cell anchor metadata from a drawing anchor."""
    marker = getattr(anchor, "_from", None)
    if marker is None and hasattr(anchor, "from_"):
        marker = getattr(anchor, "from_", None)
    if marker is not None:
        row = getattr(marker, "row", None)
        column = getattr(marker, "col", None)
        if isinstance(row, int) and isinstance(column, int):
            row_number = row + 1
            column_number = column + 1
            return {
                "anchor_row": row_number,
                "anchor_column": column_number,
                "anchor_cell": (
                    f"{get_column_letter(column_number)}{row_number}"
                ),
            }

    if isinstance(anchor, str):
        min_col, min_row, _max_col, _max_row = range_boundaries(anchor)
        return {
            "anchor_row": min_row,
            "anchor_column": min_col,
            "anchor_cell": f"{get_column_letter(min_col)}{min_row}",
        }

    return {
        "anchor_row": None,
        "anchor_column": None,
        "anchor_cell": "",
    }


def _load_reference_values(
    workbook: Any,
    cached_workbook: Any,
    reference: str,
) -> list[str]:
    """Resolve an Excel reference formula into normalized cell values."""
    if not reference:
        return []

    try:
        sheet_name, boundaries = range_to_tuple(reference)
    except ValueError:
        return []

    min_col, min_row, max_col, max_row = boundaries
    try:
        formula_sheet = workbook[sheet_name]
    except KeyError:
        return []

    cached_sheet = None
    if sheet_name in getattr(cached_workbook, "sheetnames", []):
        cached_sheet = cached_workbook[sheet_name]

    values: list[str] = []
    for row_number in range(min_row, max_row + 1):
        for column_number in range(min_col, max_col + 1):
            formula_cell = formula_sheet.cell(
                row=row_number,
                column=column_number,
            )
            raw_value = (
                cached_sheet.cell(row=row_number, column=column_number).value
                if cached_sheet is not None
                else formula_cell.value
            )
            values.append(_normalize_cell_value(raw_value))
    return values


def _extract_series_name(
    series: Any,
    workbook: Any,
    cached_workbook: Any,
    series_index: int,
) -> str:
    """Extract a human-readable chart series name."""
    text_source = getattr(series, "tx", None)
    literal_value = getattr(text_source, "v", None)
    if isinstance(literal_value, str) and literal_value.strip():
        return literal_value.strip()

    string_reference = getattr(text_source, "strRef", None)
    formula = getattr(string_reference, "f", "")
    values = _load_reference_values(workbook, cached_workbook, formula)
    for value in values:
        if value:
            return value
    return f"Series {series_index}"


def _extract_series_points(
    series: Any,
    workbook: Any,
    cached_workbook: Any,
) -> list[str]:
    """Extract paired category/value points for a chart series."""
    category_source = getattr(series, "cat", None)
    category_reference = getattr(category_source, "strRef", None)
    if category_reference is None:
        category_reference = getattr(category_source, "numRef", None)
    category_formula = getattr(category_reference, "f", "")
    value_source = getattr(series, "val", None)
    value_reference = getattr(value_source, "numRef", None)
    value_formula = getattr(value_reference, "f", "")

    categories = _load_reference_values(
        workbook,
        cached_workbook,
        category_formula,
    )
    values = _load_reference_values(workbook, cached_workbook, value_formula)
    points: list[str] = []
    for index, (category, value) in enumerate(
        zip_longest(categories, values, fillvalue=""),
        start=1,
    ):
        if not category and not value:
            continue
        label = category or f"Point {index}"
        points.append(f"{label}: {value}")
        if len(points) >= _VISUAL_SERIES_POINT_LIMIT:
            break
    return points


def _build_chart_context(
    sheet_name: str,
    region_id: str,
    chart: Any,
    workbook: Any,
    cached_workbook: Any,
) -> str:
    """Build a prompt context block for one Excel chart."""
    anchor = _extract_anchor_position(getattr(chart, "anchor", None))
    chart_type = type(chart).__name__
    chart_title = _extract_chart_title(chart) or "Untitled Chart"
    x_axis_title = _extract_title_text(getattr(chart.x_axis, "title", None))
    y_axis_title = _extract_title_text(getattr(chart.y_axis, "title", None))
    lines = [
        "## Worksheet visual context",
        f"- Sheet name: {sheet_name}",
        f"- Visual region ID: {region_id}",
        "- Visual type: chart",
        f"- Native chart type: {chart_type}",
        f"- Anchor cell: {anchor['anchor_cell'] or 'unknown'}",
        f"- Chart title: {chart_title}",
        f"- X-axis title: {x_axis_title or 'not labeled'}",
        f"- Y-axis title: {y_axis_title or 'not labeled'}",
        f"- Series count: {len(getattr(chart, 'ser', []))}",
    ]
    series_list = getattr(chart, "ser", [])
    if series_list:
        lines.extend(["", "## Series and data points"])
    for series_index, series in enumerate(series_list, start=1):
        points = _extract_series_points(series, workbook, cached_workbook)
        lines.append(f"### Series {series_index}")
        lines.append(
            "- Name: "
            + _extract_series_name(
                series,
                workbook,
                cached_workbook,
                series_index,
            )
        )
        if points:
            lines.append("- Data points:")
            lines.extend(f"  - {point}" for point in points)
        else:
            lines.append("- Data points: none recovered from workbook refs")
    return "\n".join(lines)


def _call_visual_prompt(
    llm: LLMClient,
    prompt: dict[str, Any],
    user_context: str,
) -> tuple[str, str]:
    """Call the visual-description prompt using tool-calling output."""
    messages = []
    if prompt.get("system_prompt"):
        messages.append({"role": "system", "content": prompt["system_prompt"]})
    messages.append(
        {
            "role": "user",
            "content": f"{prompt['user_prompt']}\n\n{user_context}",
        }
    )
    response = llm.call(
        messages=messages,
        stage=prompt["stage"],
        tools=prompt.get("tools"),
        tool_choice=prompt.get("tool_choice"),
    )
    return parse_vision_response(response)


def _call_visual_prompt_with_retry(
    llm: LLMClient,
    prompt: dict[str, Any],
    user_context: str,
    visual_label: str,
) -> tuple[str, str]:
    """Retry chart/image text descriptions before failing the workbook."""
    max_retries = get_xlsx_classification_max_retries()
    retry_delay = get_xlsx_classification_retry_delay()

    for attempt in range(1, max_retries + 1):
        try:
            return _call_visual_prompt(llm, prompt, user_context)
        except _RETRYABLE_CLASSIFICATION_ERRORS as exc:
            if attempt == max_retries:
                raise RuntimeError(
                    "XLSX visual description failed after "
                    f"{max_retries} attempts for {visual_label}: {exc}"
                ) from exc
            wait = retry_delay * attempt
            logger.warning(
                "XLSX visual description retry %d/%d after %.1fs for %s: %s",
                attempt,
                max_retries,
                wait,
                visual_label,
                exc,
            )
            time.sleep(wait)
    raise RuntimeError(
        "XLSX visual description exited retry loop without a response"
    )


def _describe_chart_visual(
    llm: LLMClient,
    prompt: dict[str, Any],
    sheet_name: str,
    chart: Any,
    workbook: Any,
    cached_workbook: Any,
    visual_index: int,
) -> dict[str, Any]:
    """Describe one chart using workbook metadata and tool-calling output."""
    region_id = f"visual_region_{visual_index}"
    anchor = _extract_anchor_position(getattr(chart, "anchor", None))
    page_title, content = _call_visual_prompt_with_retry(
        llm=llm,
        prompt=prompt,
        user_context=_build_chart_context(
            sheet_name=sheet_name,
            region_id=region_id,
            chart=chart,
            workbook=workbook,
            cached_workbook=cached_workbook,
        ),
        visual_label=f"chart '{sheet_name}:{region_id}'",
    )
    return {
        "region_id": region_id,
        "region_type": "visual",
        "visual_kind": "chart",
        "visual_title": _extract_chart_title(chart),
        "description_title": page_title,
        "description_content": content,
        "method": "xlsx_chart_prompt",
        **anchor,
    }


def _describe_image_visual(
    llm: LLMClient,
    prompt: dict[str, Any],
    sheet_name: str,
    image: Any,
    visual_index: int,
) -> dict[str, Any]:
    """Describe one embedded image using the shared vision flow."""
    region_id = f"visual_region_{visual_index}"
    anchor = _extract_anchor_position(getattr(image, "anchor", None))
    if not hasattr(image, "_data"):
        raise RuntimeError(
            "Embedded image bytes are unavailable for "
            f"'{sheet_name}:{region_id}'"
        )

    image_data_loader = getattr(image, "_data")
    image_bytes = image_data_loader()
    image_prompt = dict(prompt)
    image_prompt["user_prompt"] = (
        f"{prompt['user_prompt']}\n\n"
        "## Worksheet visual context\n"
        f"- Sheet name: {sheet_name}\n"
        f"- Visual region ID: {region_id}\n"
        "- Visual type: embedded image\n"
        f"- Anchor cell: {anchor['anchor_cell'] or 'unknown'}"
    )
    page_title, content = call_vision(
        llm=llm,
        img_bytes=image_bytes,
        prompt=image_prompt,
        tool_choice=image_prompt.get("tool_choice"),
    )
    return {
        "region_id": region_id,
        "region_type": "visual",
        "visual_kind": "image",
        "visual_title": "",
        "description_title": page_title,
        "description_content": content,
        "method": "xlsx_image_vision",
        **anchor,
    }


def _build_visual_regions(
    llm: LLMClient,
    prompt: dict[str, Any],
    sheet: Any,
    workbook: Any,
    cached_workbook: Any,
) -> list[dict[str, Any]]:
    """Build first-class visual regions for charts and embedded images."""
    regions: list[dict[str, Any]] = []
    visual_index = 1
    for chart in getattr(sheet, "_charts", []):
        regions.append(
            _describe_chart_visual(
                llm=llm,
                prompt=prompt,
                sheet_name=sheet.title,
                chart=chart,
                workbook=workbook,
                cached_workbook=cached_workbook,
                visual_index=visual_index,
            )
        )
        visual_index += 1
    for image in getattr(sheet, "_images", []):
        regions.append(
            _describe_image_visual(
                llm=llm,
                prompt=prompt,
                sheet_name=sheet.title,
                image=image,
                visual_index=visual_index,
            )
        )
        visual_index += 1
    return regions


def _is_chartsheet(sheet: Any) -> bool:
    """Return True when the workbook tab is a dedicated chartsheet."""
    return type(sheet).__name__ == "Chartsheet"


def _build_visual_metadata(sheet: Any) -> dict[str, Any]:
    """Summarize non-grid visual content attached to a sheet."""
    charts = getattr(sheet, "_charts", [])
    images = getattr(sheet, "_images", [])
    chart_titles = []
    for chart in charts:
        title = _extract_chart_title(chart)
        chart_type = type(chart).__name__
        chart_titles.append(f"{chart_type}: {title}" if title else chart_type)
    return {
        "sheet_kind": ("chartsheet" if _is_chartsheet(sheet) else "worksheet"),
        "chart_count": len(charts),
        "image_count": len(images),
        "chart_titles": chart_titles,
    }


def _build_cached_values(
    cached_sheet: Worksheet,
) -> dict[tuple[int, int], Any]:
    """Build a lookup of cached display values from a data_only worksheet."""
    values: dict[tuple[int, int], Any] = {}
    for row in cached_sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                values[(cell.row, cell.column)] = cell.value
    return values


def _collect_populated_cells(
    sheet: Worksheet,
    cached_values: dict[tuple[int, int], Any],
) -> dict[str, Any] | None:
    """Collect non-empty worksheet cells and compact grid statistics."""
    min_row = 0
    min_col = 0
    max_row = 0
    max_col = 0
    rows: dict[int, dict[int, str]] = {}
    populated_columns: set[int] = set()
    non_empty_cells = 0
    formula_cells = 0
    max_populated_cells_in_row = 0

    for row in sheet.iter_rows():
        populated_cells_in_row: dict[int, str] = {}
        for cell in row:
            value = _normalize_cell_value(
                cached_values.get((cell.row, cell.column), cell.value)
            )
            if not value:
                continue
            if min_row == 0 or cell.row < min_row:
                min_row = cell.row
            if min_col == 0 or cell.column < min_col:
                min_col = cell.column
            max_row = max(max_row, cell.row)
            max_col = max(max_col, cell.column)
            populated_cells_in_row[cell.column] = value
            populated_columns.add(cell.column)
            non_empty_cells += 1
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells += 1
        if populated_cells_in_row:
            rows[row[0].row] = populated_cells_in_row
            max_populated_cells_in_row = max(
                max_populated_cells_in_row, len(populated_cells_in_row)
            )

    if min_row == 0:
        return None
    return {
        "bounds": (min_row, min_col, max_row, max_col),
        "rows": rows,
        "columns": sorted(populated_columns),
        "row_count": len(rows),
        "column_count": len(populated_columns),
        "row_span": max_row - min_row + 1,
        "column_span": max_col - min_col + 1,
        "blank_rows_omitted": (max_row - min_row + 1) - len(rows),
        "blank_columns_omitted": (max_col - min_col + 1)
        - len(populated_columns),
        "non_empty_cells": non_empty_cells,
        "formula_cells": formula_cells,
        "max_populated_cells_in_row": max_populated_cells_in_row,
    }


def _build_used_range(bounds: tuple[int, int, int, int]) -> str:
    """Build Excel A1 range notation from populated bounds."""
    min_row, min_col, max_row, max_col = bounds
    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    if start == end:
        return start
    return f"{start}:{end}"


def _truncate_region_cell(value: str) -> str:
    """Truncate and escape region preview cell text."""
    if len(value) > _REGION_PREVIEW_MAX_VALUE_LENGTH:
        value = value[: _REGION_PREVIEW_MAX_VALUE_LENGTH - 3] + "..."
    return _escape_table_text(value)


def _serialize_region_summary(region: SheetRegion) -> dict[str, Any]:
    """Serialize one detected region into JSON-safe summary metadata."""
    return {
        "region_id": region.region_id,
        "region_type": region.region_type,
        "used_range": build_region_used_range(region),
        "row_count": region.row_count,
        "column_count": region.column_count,
        "non_empty_cells": region.non_empty_cells,
        "dense_score": region.dense_score,
        "native_table_name": region.native_table_name,
    }


def _serialize_region_detail(region: SheetRegion) -> dict[str, Any]:
    """Serialize one detected region with row/cell payloads."""
    rows: list[dict[str, Any]] = []
    for row_number, row_values in zip(region.row_numbers, region.rows):
        rows.append(
            {
                "row_number": row_number,
                "cells": [
                    {
                        "column_number": column_number,
                        "value": row_values[column_number],
                    }
                    for column_number in sorted(row_values)
                ],
            }
        )

    detail = _serialize_region_summary(region)
    detail.update(
        {
            "row_numbers": region.row_numbers,
            "column_numbers": region.column_numbers,
            "rows": rows,
            "metadata": dict(region.metadata),
        }
    )
    return detail


def _visual_overlaps_region(
    visual_region: dict[str, Any],
    region: SheetRegion,
) -> bool:
    """Check whether a visual anchor falls within a grid region."""
    anchor_row = visual_region.get("anchor_row")
    anchor_column = visual_region.get("anchor_column")
    if not isinstance(anchor_row, int) or not isinstance(anchor_column, int):
        return False
    return (
        region.min_row <= anchor_row <= region.max_row
        and region.min_col <= anchor_column <= region.max_col
    )


def _region_sample_text(region: SheetRegion) -> str:
    """Build a short plain-text sample from the first populated region row."""
    for row in region.rows:
        values = [value for value in row.values() if value]
        if not values:
            continue
        sample = " | ".join(values[:3])
        if len(sample) > _REGION_PREVIEW_MAX_VALUE_LENGTH:
            return sample[: _REGION_PREVIEW_MAX_VALUE_LENGTH - 3] + "..."
        return sample
    return "n/a"


def _build_region_preview(region: SheetRegion | None) -> str:
    """Build a compact preview for a detected region."""
    if region is None:
        return ""

    displayed_columns = region.column_numbers[:_REGION_PREVIEW_MAX_COLUMNS]
    header_cells = [
        get_column_letter(column_number) for column_number in displayed_columns
    ]
    lines = [
        f"- Region ID: {region.region_id}",
        f"- Region type: {region.region_type}",
        f"- Used range: {build_region_used_range(region)}",
        (
            f"- Region span: rows={region.row_count}, "
            f"columns={region.column_count}"
        ),
        f"- Dense-table score: {region.dense_score}",
        "",
        f"| Row | {' | '.join(header_cells)} |",
        f"| {' | '.join(['---'] * (len(header_cells) + 1))} |",
    ]
    for row_number, row_values in zip(
        region.row_numbers[:_REGION_PREVIEW_MAX_ROWS],
        region.rows[:_REGION_PREVIEW_MAX_ROWS],
    ):
        rendered_row = [
            _truncate_region_cell(row_values.get(column_number, ""))
            for column_number in displayed_columns
        ]
        lines.append(f"| {row_number} | {' | '.join(rendered_row)} |")

    if (
        len(region.row_numbers) > _REGION_PREVIEW_MAX_ROWS
        or len(region.column_numbers) > _REGION_PREVIEW_MAX_COLUMNS
    ):
        lines.append("... omitted candidate region rows or columns ...")
    return "\n".join(lines)


def _build_framing_summary(
    regions: list[SheetRegion],
    dense_region: SheetRegion | None,
) -> str:
    """Summarize non-dense framing regions for classification context."""
    framing_regions = [
        region
        for region in regions
        if dense_region is None or region.region_id != dense_region.region_id
    ]
    if not framing_regions:
        return "No separate framing regions detected."

    lines = []
    for region in framing_regions:
        lines.append(
            (
                f"- {region.region_id}: {region.region_type}, "
                f"range={build_region_used_range(region)}, "
                f"rows={region.row_count}, columns={region.column_count}, "
                f"sample={_region_sample_text(region)}"
            )
        )
    return "\n".join(lines)


def _build_region_metadata(
    sheet: Worksheet,
    cached_values: dict[tuple[int, int], Any],
    visual_regions: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    """Build region-level worksheet metadata for classifier prompts."""
    visual_regions = [dict(region) for region in (visual_regions or [])]
    regions = build_sheet_regions(sheet, cached_values)
    if not regions:
        metadata = _default_region_metadata(visual_regions)
        metadata["typed_regions"] = list(visual_regions)
        metadata["total_region_count"] = len(visual_regions)
        return metadata

    dense_region = select_dense_table_region(regions)
    framing_regions = []
    typed_regions: list[dict[str, Any]] = []
    dense_table_regions: list[dict[str, Any]] = []
    small_table_regions: list[dict[str, Any]] = []
    mixed_regions: list[dict[str, Any]] = []

    for region in regions:
        detail = _serialize_region_detail(region)
        linked_visual_region_ids = [
            visual_region["region_id"]
            for visual_region in visual_regions
            if _visual_overlaps_region(visual_region, region)
        ]
        detail["source_region_type"] = region.region_type
        detail["linked_visual_region_ids"] = linked_visual_region_ids

        if region.region_type in ("native_table", "dense_table_candidate"):
            dense_detail = dict(detail)
            dense_detail["region_type"] = "dense_table"
            dense_detail["typed_region_type"] = (
                "mixed" if linked_visual_region_ids else "dense_table"
            )
            dense_table_regions.append(dense_detail)

        if linked_visual_region_ids:
            typed_detail = dict(detail)
            typed_detail["region_type"] = "mixed"
            mixed_regions.append(typed_detail)
        elif region.region_type in ("native_table", "dense_table_candidate"):
            typed_detail = dict(detail)
            typed_detail["region_type"] = "dense_table"
        elif region.row_count >= 2 and region.column_count >= 2:
            typed_detail = dict(detail)
            typed_detail["region_type"] = "small_table"
            small_table_regions.append(dict(typed_detail))
        else:
            typed_detail = dict(detail)
            typed_detail["region_type"] = "framing"
            framing_regions.append(
                {
                    **_serialize_region_summary(region),
                    "region_type": "framing",
                }
            )
        typed_regions.append(typed_detail)

    for visual_region in visual_regions:
        visual_region["linked_grid_region_ids"] = [
            region.region_id
            for region in regions
            if _visual_overlaps_region(visual_region, region)
        ]

    selected_region_detail = None
    if dense_region is not None:
        selected_region_detail = _serialize_region_detail(dense_region)
        selected_region_detail["linked_visual_region_ids"] = [
            visual_region["region_id"]
            for visual_region in visual_regions
            if _visual_overlaps_region(visual_region, dense_region)
        ]
        selected_region_detail["typed_region_type"] = (
            "mixed"
            if selected_region_detail["linked_visual_region_ids"]
            else "dense_table"
        )
    return {
        "region_count": len(regions),
        "visual_region_count": len(visual_regions),
        "total_region_count": len(typed_regions) + len(visual_regions),
        "sheet_regions": [
            _serialize_region_summary(region) for region in regions
        ],
        "typed_regions": typed_regions + visual_regions,
        "dense_table_regions": dense_table_regions,
        "small_table_regions": small_table_regions,
        "mixed_regions": mixed_regions,
        "visual_regions": visual_regions,
        "dense_table_region": selected_region_detail,
        "framing_regions": framing_regions,
        "dense_table_used_range": (
            build_region_used_range(dense_region)
            if dense_region is not None
            else ""
        ),
        "dense_table_region_preview": _build_region_preview(dense_region),
        "framing_summary": _build_framing_summary(regions, dense_region),
        "dense_table_candidate_detected": dense_region is not None,
    }


def _build_visual_lines(visual_metadata: dict[str, Any]) -> list[str]:
    """Build markdown lines describing charts and images on a sheet."""
    if (
        visual_metadata["chart_count"] == 0
        and visual_metadata["image_count"] == 0
    ):
        return []

    lines = [
        "",
        "## Visual Elements",
        f"- Charts: {visual_metadata['chart_count']}",
        f"- Images: {visual_metadata['image_count']}",
    ]
    for index, chart_title in enumerate(
        visual_metadata["chart_titles"], start=1
    ):
        lines.append(f"- Chart {index}: {chart_title}")
    return lines


def _build_visual_only_sheet_content(
    sheet_name: str,
    visual_metadata: dict[str, Any],
) -> str:
    """Serialize a visual-only sheet into a compact markdown summary."""
    lines = [
        f"# Sheet: {sheet_name}",
        f"- Sheet type: {visual_metadata['sheet_kind']}",
        "- Content type: visual_only",
    ]
    lines.extend(_build_visual_lines(visual_metadata))
    return "\n".join(lines)


def _build_visual_only_sheet_metadata(
    sheet_name: str,
    visual_metadata: dict[str, Any],
) -> dict[str, Any]:
    """Build metadata for sheets containing visuals but no cell grid."""
    metadata = _build_empty_sheet_metadata(sheet_name, visual_metadata)
    metadata["content_kind"] = "visual_only"
    return metadata


def _default_region_metadata(
    visual_regions: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    """Build default region metadata for sheets without grid analysis."""
    visual_regions = [dict(region) for region in (visual_regions or [])]
    return {
        "region_count": 0,
        "visual_region_count": len(visual_regions),
        "total_region_count": len(visual_regions),
        "sheet_regions": [],
        "typed_regions": list(visual_regions),
        "dense_table_regions": [],
        "small_table_regions": [],
        "mixed_regions": [],
        "visual_regions": visual_regions,
        "dense_table_region": None,
        "framing_regions": [],
        "dense_table_used_range": "",
        "dense_table_region_preview": "",
        "framing_summary": "",
        "dense_table_candidate_detected": False,
    }


def _append_visual_descriptions(
    content: str,
    visual_regions: list[dict[str, Any]],
) -> str:
    """Append rich visual descriptions to serialized sheet content."""
    if not visual_regions:
        return content

    lines = [content.rstrip(), "", "## Visual Region Descriptions"]
    for visual_region in visual_regions:
        lines.append(f"### {visual_region['region_id']}")
        lines.append(f"- Visual type: {visual_region['visual_kind']}")
        if visual_region.get("anchor_cell"):
            lines.append(f"- Anchor cell: {visual_region['anchor_cell']}")
        if visual_region.get("visual_title"):
            lines.append(f"- Native title: {visual_region['visual_title']}")
        if visual_region.get("linked_grid_region_ids"):
            lines.append(
                "- Linked grid regions: "
                + ", ".join(visual_region["linked_grid_region_ids"])
            )
        lines.extend(["", visual_region["description_content"].strip(), ""])
    return "\n".join(line for line in lines if line is not None)


def _serialize_sheet(
    sheet: Any,
    cached_values: dict[tuple[int, int], Any],
) -> tuple[str, dict[str, Any]]:
    """Serialize a worksheet or chartsheet into markdown-like text."""
    visual_metadata = _build_visual_metadata(sheet)
    if _is_chartsheet(sheet):
        if (
            visual_metadata["chart_count"] == 0
            and visual_metadata["image_count"] == 0
        ):
            return "## Empty Sheet", _build_empty_sheet_metadata(
                sheet.title, visual_metadata
            )
        return (
            _build_visual_only_sheet_content(sheet.title, visual_metadata),
            _build_visual_only_sheet_metadata(sheet.title, visual_metadata),
        )

    cell_data = _collect_populated_cells(sheet, cached_values)
    if cell_data is None:
        if (
            visual_metadata["chart_count"] > 0
            or visual_metadata["image_count"] > 0
        ):
            return (
                _build_visual_only_sheet_content(sheet.title, visual_metadata),
                _build_visual_only_sheet_metadata(
                    sheet.title, visual_metadata
                ),
            )
        return "## Empty Sheet", _build_empty_sheet_metadata(
            sheet.title, visual_metadata
        )

    used_range = _build_used_range(cell_data["bounds"])
    lines = _build_sheet_header_lines(
        sheet_name=sheet.title,
        used_range=used_range,
        cell_data=cell_data,
        visual_metadata=visual_metadata,
    )
    _append_sheet_rows(
        lines=lines,
        rows=cell_data["rows"],
        column_numbers=cell_data["columns"],
    )
    lines.extend(_build_visual_lines(visual_metadata))

    metadata = _build_base_sheet_metadata(sheet.title, visual_metadata)
    metadata.update(
        {
            "content_kind": "grid",
            "used_range": used_range,
            "row_count": cell_data["row_count"],
            "column_count": cell_data["column_count"],
            "row_span": cell_data["row_span"],
            "column_span": cell_data["column_span"],
            "blank_rows_omitted": cell_data["blank_rows_omitted"],
            "blank_columns_omitted": cell_data["blank_columns_omitted"],
            "non_empty_cells": cell_data["non_empty_cells"],
            "formula_cells": cell_data["formula_cells"],
            "merged_range_count": len(sheet.merged_cells.ranges),
            "max_populated_cells_in_row": cell_data[
                "max_populated_cells_in_row"
            ],
        }
    )
    return "\n".join(lines), metadata


def _build_sheet_header_lines(
    sheet_name: str,
    used_range: str,
    cell_data: dict[str, Any],
    visual_metadata: dict[str, Any],
) -> list[str]:
    """Build the static header rows for a serialized worksheet."""
    header_cells = [
        get_column_letter(column) for column in cell_data["columns"]
    ]
    return [
        f"# Sheet: {sheet_name}",
        f"- Sheet type: {visual_metadata['sheet_kind']}",
        f"- Used range: {used_range}",
        (
            f"- Populated grid: rows={cell_data['row_count']}, "
            f"columns={cell_data['column_count']}"
        ),
        (
            f"- Used span: rows={cell_data['row_span']}, "
            f"columns={cell_data['column_span']}"
        ),
        (
            f"- Omitted blank space: rows={cell_data['blank_rows_omitted']}, "
            f"columns={cell_data['blank_columns_omitted']}"
        ),
        (
            f"- Visual counts: charts={visual_metadata['chart_count']}, "
            f"images={visual_metadata['image_count']}"
        ),
        "",
        f"| Row | {' | '.join(header_cells)} |",
        f"| {' | '.join(['---'] * (len(header_cells) + 1))} |",
    ]


def _append_sheet_rows(
    lines: list[str],
    rows: dict[int, dict[int, str]],
    column_numbers: list[int],
) -> None:
    """Append compact worksheet rows for the populated grid only."""
    for row_number in sorted(rows):
        rendered_row = [
            _escape_table_text(rows[row_number].get(column, ""))
            for column in column_numbers
        ]
        lines.append(f"| {row_number} | {' | '.join(rendered_row)} |")


def _truncate_line(line: str) -> str:
    """Truncate a line that exceeds the preview width limit."""
    if len(line) <= _PREVIEW_MAX_LINE_LENGTH:
        return line
    return line[:_PREVIEW_MAX_LINE_LENGTH] + " ..."


def _build_preview(sheet_content: str) -> str:
    """Build a compact preview for classification prompts."""
    lines = sheet_content.splitlines()
    if len(lines) <= _PREVIEW_HEAD_LINES + _PREVIEW_TAIL_LINES:
        return "\n".join(_truncate_line(line) for line in lines)

    preview_lines = [
        _truncate_line(line) for line in lines[:_PREVIEW_HEAD_LINES]
    ]
    preview_lines.append("... omitted sheet rows ...")
    preview_lines.extend(
        _truncate_line(line) for line in lines[-_PREVIEW_TAIL_LINES:]
    )
    return "\n".join(preview_lines)


def _build_classifier_message(
    prompt: dict[str, Any],
    sheet_content: str,
    sheet_metadata: dict[str, Any],
) -> str:
    """Build the XLSX classification user message."""
    region_lines = [
        "## Local layout analysis",
        f"- Detected regions: {sheet_metadata['region_count']}",
        (
            "- Visual regions detected locally: "
            f"{sheet_metadata.get('visual_region_count', 0)}"
        ),
        "- Dense-table candidate detected locally: "
        + (
            "yes" if sheet_metadata["dense_table_candidate_detected"] else "no"
        ),
    ]
    for region in sheet_metadata["sheet_regions"]:
        region_lines.append(
            (
                f"- {region['region_id']}: {region['region_type']}, "
                f"range={region['used_range']}, "
                f"rows={region['row_count']}, "
                f"columns={region['column_count']}, "
                f"score={region['dense_score']}"
            )
        )
    if sheet_metadata["dense_table_used_range"]:
        region_lines.append(
            "- Selected dense-table region: "
            f"{sheet_metadata['dense_table_used_range']}"
        )
    for visual_region in sheet_metadata.get("visual_regions", []):
        visual_title = (
            visual_region.get("visual_title")
            or visual_region.get("description_title")
            or "Untitled visual"
        )
        region_lines.append(
            (
                f"- {visual_region['region_id']}: "
                f"{visual_region['visual_kind']}, "
                f"anchor={visual_region.get('anchor_cell', '') or 'unknown'}, "
                f"title={visual_title}"
            )
        )

    if sheet_metadata["dense_table_candidate_detected"]:
        preview_section = (
            "## Candidate dense-table region\n"
            f"{sheet_metadata['dense_table_region_preview']}\n\n"
            "## Framing outside candidate region\n"
            f"{sheet_metadata['framing_summary']}"
        )
    else:
        preview_section = (
            "## Framing regions\n"
            f"{sheet_metadata['framing_summary']}\n\n"
            "## Sheet preview\n"
            f"{_build_preview(sheet_content)}"
        )

    return (
        f"{prompt['user_prompt']}\n\n"
        "## Sheet context\n"
        f"- Sheet name: {sheet_metadata['sheet_name']}\n"
        f"- Sheet type: {sheet_metadata['sheet_kind']}\n"
        f"- Used range: {sheet_metadata['used_range']}\n"
        "- Populated grid: "
        f"rows={sheet_metadata['row_count']}, "
        f"columns={sheet_metadata['column_count']}\n"
        "- Used span: "
        f"rows={sheet_metadata['row_span']}, "
        f"columns={sheet_metadata['column_span']}\n"
        "- Blank space omitted: "
        f"rows={sheet_metadata['blank_rows_omitted']}, "
        f"columns={sheet_metadata['blank_columns_omitted']}\n"
        f"- Merged ranges: {sheet_metadata['merged_range_count']}\n"
        f"- Charts: {sheet_metadata['chart_count']}\n"
        f"- Images: {sheet_metadata['image_count']}\n"
        f"- Estimated sheet tokens: {sheet_metadata['estimated_tokens']}\n"
        f"- Inline sheet token limit: {sheet_metadata['token_limit']}\n\n"
        + "\n".join(region_lines)
        + "\n\n"
        + preview_section
    )


def _get_encoder(model: str):
    """Get a tokenizer for the configured model, with a safe fallback."""
    try:
        return tiktoken.encoding_for_model(model)
    except KeyError:
        return tiktoken.get_encoding("o200k_base")


def _count_tokens(text: str, model: str) -> int:
    """Count tokens for serialized sheet text.

    Params: text, model. Returns: int.
    """
    return len(_get_encoder(model).encode(text))


def _parse_classification_response(response: dict[str, Any]) -> dict[str, Any]:
    """Parse the XLSX sheet classification tool response."""
    parsed = parse_tool_arguments(response)

    handling_mode = parsed.get("handling_mode")
    contains_dense_table = parsed.get("contains_dense_table")
    confidence = parsed.get("confidence")
    rationale = parsed.get("rationale")
    if not isinstance(handling_mode, str) or not handling_mode:
        if isinstance(contains_dense_table, bool):
            handling_mode = (
                "dense_table_candidate"
                if contains_dense_table
                else "page_like"
            )
        else:
            raise ValueError(
                "LLM tool arguments missing handling_mode "
                "and contains_dense_table fallback"
            )
    if handling_mode not in ("page_like", "dense_table_candidate"):
        raise ValueError(
            "LLM tool arguments contain invalid "
            f"handling_mode: '{handling_mode}'"
        )
    if not isinstance(confidence, (int, float)):
        raise ValueError("LLM tool arguments missing numeric confidence")
    conf_value = float(confidence)
    if conf_value < 0.0 or conf_value > 1.0:
        raise ValueError(
            f"LLM tool arguments confidence out of range: {confidence}"
        )
    if not isinstance(rationale, str) or not rationale:
        raise ValueError("LLM tool arguments missing non-empty rationale")

    return {
        "handling_mode": handling_mode,
        "contains_dense_table": handling_mode == "dense_table_candidate",
        "confidence": float(confidence),
        "rationale": rationale,
        "classification": handling_mode,
    }


def _classify_sheet(
    llm: LLMClient,
    prompt: dict[str, Any],
    sheet_content: str,
    sheet_metadata: dict[str, Any],
) -> dict[str, Any]:
    """Run the standalone XLSX classifier against one sheet."""
    messages = []
    if prompt.get("system_prompt"):
        messages.append({"role": "system", "content": prompt["system_prompt"]})
    messages.append(
        {
            "role": "user",
            "content": _build_classifier_message(
                prompt=prompt,
                sheet_content=sheet_content,
                sheet_metadata=sheet_metadata,
            ),
        }
    )
    response = llm.call(
        messages=messages,
        stage=prompt["stage"],
        tools=prompt.get("tools"),
        tool_choice=prompt.get("tool_choice"),
    )
    return _parse_classification_response(response)


def _classify_sheet_with_retry(
    llm: LLMClient,
    prompt: dict[str, Any],
    sheet_content: str,
    sheet_metadata: dict[str, Any],
) -> dict[str, Any]:
    """Retry XLSX sheet classification before failing the workbook."""
    max_retries = get_xlsx_classification_max_retries()
    retry_delay = get_xlsx_classification_retry_delay()

    for attempt in range(1, max_retries + 1):
        try:
            return _classify_sheet(
                llm=llm,
                prompt=prompt,
                sheet_content=sheet_content,
                sheet_metadata=sheet_metadata,
            )
        except _RETRYABLE_CLASSIFICATION_ERRORS as exc:
            if attempt == max_retries:
                raise RuntimeError(
                    "XLSX sheet classification failed after "
                    f"{max_retries} attempts for sheet "
                    f"'{sheet_metadata['sheet_name']}': {exc}"
                ) from exc
            wait = retry_delay * attempt
            logger.warning(
                "XLSX classification retry %d/%d after %.1fs for sheet "
                "'%s': %s",
                attempt,
                max_retries,
                wait,
                sheet_metadata["sheet_name"],
                exc,
            )
            time.sleep(wait)
    raise RuntimeError(
        "XLSX classification exited retry loop without a response"
    )


def _build_sheet_result(
    page_number: int,
    sheet: Any,
    content: str,
    metadata: dict[str, Any],
) -> PageResult:
    """Build the standard page result for a processed workbook tab."""
    return PageResult(
        page_number=page_number,
        page_title=sheet.title,
        content=content,
        method="xlsx_sheet_classification",
        metadata=metadata,
    )


def _open_workbooks(file_path: str) -> tuple[Any, Any]:
    """Open formula and cached-value workbooks from the same file."""
    name = Path(file_path).name
    try:
        workbook: Any = load_workbook(filename=file_path, data_only=False)
    except (OSError, ValueError) as exc:
        raise RuntimeError(f"Failed to open XLSX '{name}': {exc}") from exc
    try:
        cached_workbook: Any = load_workbook(
            filename=file_path, data_only=True
        )
    except (OSError, ValueError) as exc:
        workbook.close()
        raise RuntimeError(f"Failed to open XLSX '{name}': {exc}") from exc
    return workbook, cached_workbook


def process_xlsx(file_path: str, llm: LLMClient) -> ExtractionResult:
    """Classify each worksheet for inline handling vs dense-table follow-up.

    Params:
        file_path: Absolute path to the XLSX file
        llm: LLMClient instance

    Returns:
        ExtractionResult with one PageResult per worksheet

    Example:
        >>> result = process_xlsx("/data/report.xlsx", llm)
        >>> result.filetype
        "xlsx"
    """
    prompt = load_prompt("xlsx_sheet_classification")
    model = get_stage_model_config(prompt["stage"])["model"]
    token_limit = get_xlsx_sheet_token_limit()
    workbook, cached_workbook = _open_workbooks(file_path)
    visual_prompt: dict[str, Any] | None = None

    sheets = [workbook[sheet_name] for sheet_name in workbook.sheetnames]
    pages: list[PageResult] = []

    try:
        for page_number, sheet in enumerate(sheets, start=1):
            if _is_chartsheet(sheet):
                cached_values: dict[tuple[int, int], Any] = {}
            else:
                cached_values = _build_cached_values(
                    cached_workbook[sheet.title]
                )
            content, metadata = _serialize_sheet(sheet, cached_values)
            visual_regions: list[dict[str, Any]] = []
            if metadata["chart_count"] > 0 or metadata["image_count"] > 0:
                if visual_prompt is None:
                    visual_prompt = load_prompt(
                        "xlsx_visual_extraction_vision"
                    )
                visual_regions = _build_visual_regions(
                    llm=llm,
                    prompt=visual_prompt,
                    sheet=sheet,
                    workbook=workbook,
                    cached_workbook=cached_workbook,
                )
            if metadata["content_kind"] == "grid":
                metadata.update(
                    _build_region_metadata(
                        sheet,
                        cached_values,
                        visual_regions,
                    )
                )
            else:
                metadata.update(_default_region_metadata(visual_regions))
            content = _append_visual_descriptions(
                content,
                metadata["visual_regions"],
            )
            metadata["sheet_layout_kind"] = (
                "mixed"
                if (
                    metadata["content_kind"] == "grid"
                    and metadata["visual_region_count"] > 0
                )
                else metadata["content_kind"]
            )
            metadata["estimated_tokens"] = 0
            metadata["token_limit"] = token_limit
            metadata["threshold_exceeded"] = False

            if metadata["content_kind"] == "empty":
                metadata["handling_mode"] = "page_like"
                metadata["classification"] = "empty_sheet"
                metadata["contains_dense_table"] = False
                metadata["confidence"] = 1.0
                metadata["rationale"] = "Sheet contains no populated cells."
                pages.append(
                    _build_sheet_result(page_number, sheet, content, metadata)
                )
                continue

            if metadata["content_kind"] == "visual_only":
                metadata["handling_mode"] = "page_like"
                metadata["classification"] = "page_like"
                metadata["contains_dense_table"] = False
                metadata["confidence"] = 1.0
                metadata["rationale"] = (
                    "Sheet contains visual elements without a dense cell grid."
                )
                pages.append(
                    _build_sheet_result(page_number, sheet, content, metadata)
                )
                continue

            token_count = _count_tokens(content, model)
            metadata["estimated_tokens"] = token_count
            metadata["threshold_exceeded"] = token_count > token_limit
            metadata.update(
                _classify_sheet_with_retry(
                    llm=llm,
                    prompt=prompt,
                    sheet_content=content,
                    sheet_metadata=metadata,
                )
            )
            pages.append(
                _build_sheet_result(page_number, sheet, content, metadata)
            )
    finally:
        workbook.close()
        cached_workbook.close()

    return ExtractionResult(
        file_path=file_path,
        filetype="xlsx",
        pages=pages,
        total_pages=len(pages),
        pages_succeeded=len(pages),
        pages_failed=0,
    )
