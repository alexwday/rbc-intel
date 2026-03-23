"""Local layout detection for XLSX sheets.

Identifies populated regions before worksheet compaction so later
stages can distinguish dense table candidates from surrounding notes
or framing content.
"""

from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any

from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class SheetCell:
    """One populated worksheet cell with preserved coordinates.

    Params:
        row: 1-based worksheet row number
        column: 1-based worksheet column number
        value: Normalized display value
        is_formula: Whether the source cell contains a formula

    Example:
        >>> cell = SheetCell(2, 3, "100", False)
        >>> cell.column
        3
    """

    row: int
    column: int
    value: str
    is_formula: bool


@dataclass
class SheetRegion:
    """A contiguous worksheet region detected before serialization.

    Params:
        region_id: Stable ID for the region within the sheet
        region_type: "native_table", "dense_table_candidate", or "framing"
        min_row: Top row of the region bounds
        min_col: Left column of the region bounds
        max_row: Bottom row of the region bounds
        max_col: Right column of the region bounds
        non_empty_cells: Number of populated cells inside the region
        row_count: Number of rows spanned by the region
        column_count: Number of columns spanned by the region
        row_numbers: Sorted row numbers present in the region span
        column_numbers: Sorted column numbers present in the region span
        rows: Per-row cell mappings keyed by column number
        dense_score: Heuristic dense-table score from 0 to 1
        native_table_name: Excel table name when sourced from a native table

    Example:
        >>> region = SheetRegion(
        ...     region_id="region_1",
        ...     region_type="framing",
        ...     min_row=1,
        ...     min_col=1,
        ...     max_row=1,
        ...     max_col=2,
        ...     non_empty_cells=2,
        ...     row_count=1,
        ...     column_count=2,
        ...     row_numbers=[1],
        ...     column_numbers=[1, 2],
        ...     rows=[{1: "Title", 2: "Notes"}],
        ... )
        >>> region.region_id
        'region_1'
    """

    region_id: str
    region_type: str
    min_row: int
    min_col: int
    max_row: int
    max_col: int
    non_empty_cells: int
    row_count: int
    column_count: int
    row_numbers: list[int]
    column_numbers: list[int]
    rows: list[dict[int, str]]
    dense_score: float = 0.0
    native_table_name: str = ""
    metadata: dict[str, Any] = field(default_factory=dict)


def _normalize_cell_value(value: Any) -> str:
    """Normalize a worksheet cell value into stable text."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0:
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value).replace("\r\n", "\n").replace("\r", "\n")


def _split_cells_by_blank_rows(
    cells: list[SheetCell],
) -> list[list[SheetCell]] | None:
    """Split cells into row bands separated by fully blank rows."""
    if not cells:
        return None
    rows_present = {cell.row for cell in cells}
    min_row = min(rows_present)
    max_row = max(rows_present)
    row_ranges: list[tuple[int, int]] = []
    start: int | None = None
    for row_number in range(min_row, max_row + 1):
        if row_number in rows_present:
            if start is None:
                start = row_number
            continue
        if start is not None:
            row_ranges.append((start, row_number - 1))
            start = None
    if start is not None:
        row_ranges.append((start, max_row))
    if len(row_ranges) <= 1:
        return None
    return [
        [cell for cell in cells if start_row <= cell.row <= end_row]
        for start_row, end_row in row_ranges
    ]


def _split_cells_by_blank_columns(
    cells: list[SheetCell],
) -> list[list[SheetCell]] | None:
    """Split cells into column bands separated by fully blank columns."""
    if not cells:
        return None
    columns_present = {cell.column for cell in cells}
    min_col = min(columns_present)
    max_col = max(columns_present)
    column_ranges: list[tuple[int, int]] = []
    start: int | None = None
    for column_number in range(min_col, max_col + 1):
        if column_number in columns_present:
            if start is None:
                start = column_number
            continue
        if start is not None:
            column_ranges.append((start, column_number - 1))
            start = None
    if start is not None:
        column_ranges.append((start, max_col))
    if len(column_ranges) <= 1:
        return None
    return [
        [cell for cell in cells if start_col <= cell.column <= end_col]
        for start_col, end_col in column_ranges
    ]


def _split_into_leaf_groups(cells: list[SheetCell]) -> list[list[SheetCell]]:
    """Recursively split cells across blank row and column bands."""
    pending = [cells]
    leaves: list[list[SheetCell]] = []
    while pending:
        current = pending.pop()
        row_groups = _split_cells_by_blank_rows(current)
        if row_groups:
            pending.extend(reversed(row_groups))
            continue
        column_groups = _split_cells_by_blank_columns(current)
        if column_groups:
            pending.extend(reversed(column_groups))
            continue
        leaves.append(current)
    return leaves


def _build_rows(
    cells: list[SheetCell],
) -> tuple[list[int], list[dict[int, str]]]:
    """Build ordered row mappings from a cell collection."""
    rows_by_number: dict[int, dict[int, str]] = {}
    for cell in sorted(cells, key=lambda item: (item.row, item.column)):
        rows_by_number.setdefault(cell.row, {})[cell.column] = cell.value
    row_numbers = sorted(rows_by_number)
    rows = [rows_by_number[row_number] for row_number in row_numbers]
    return row_numbers, rows


def _build_region(
    cells: list[SheetCell],
    region_type: str = "framing",
    native_table_name: str = "",
) -> SheetRegion:
    """Build a SheetRegion from populated cells."""
    min_row = min(cell.row for cell in cells)
    min_col = min(cell.column for cell in cells)
    max_row = max(cell.row for cell in cells)
    max_col = max(cell.column for cell in cells)
    row_numbers, rows = _build_rows(cells)
    column_numbers = list(range(min_col, max_col + 1))
    return SheetRegion(
        region_id="",
        region_type=region_type,
        min_row=min_row,
        min_col=min_col,
        max_row=max_row,
        max_col=max_col,
        non_empty_cells=len(cells),
        row_count=max_row - min_row + 1,
        column_count=max_col - min_col + 1,
        row_numbers=row_numbers,
        column_numbers=column_numbers,
        rows=rows,
        native_table_name=native_table_name,
    )


def _iter_native_tables(sheet: Worksheet) -> list[Table]:
    """Return native Excel tables attached to a worksheet."""
    tables = getattr(sheet, "tables", {})
    return list(tables.values())


def _cell_value_from_sheet(
    sheet: Worksheet,
    cached_values: dict[tuple[int, int], Any],
    row_number: int,
    column_number: int,
) -> tuple[str, bool]:
    """Get normalized cell text and formula status from a worksheet."""
    source_cell = sheet.cell(row=row_number, column=column_number)
    raw_value = cached_values.get(
        (row_number, column_number), source_cell.value
    )
    return _normalize_cell_value(raw_value), (
        isinstance(source_cell.value, str)
        and source_cell.value.startswith("=")
    )


def _build_native_region(
    sheet: Worksheet,
    cached_values: dict[tuple[int, int], Any],
    table: Table,
) -> SheetRegion:
    """Build a SheetRegion from an Excel native table reference."""
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    row_numbers = list(range(min_row, max_row + 1))
    column_numbers = list(range(min_col, max_col + 1))
    rows: list[dict[int, str]] = []
    non_empty_cells = 0
    for row_number in row_numbers:
        row_data: dict[int, str] = {}
        for column_number in column_numbers:
            value, _ = _cell_value_from_sheet(
                sheet=sheet,
                cached_values=cached_values,
                row_number=row_number,
                column_number=column_number,
            )
            if not value:
                continue
            row_data[column_number] = value
            non_empty_cells += 1
        rows.append(row_data)
    return SheetRegion(
        region_id="",
        region_type="native_table",
        min_row=min_row,
        min_col=min_col,
        max_row=max_row,
        max_col=max_col,
        non_empty_cells=non_empty_cells,
        row_count=max_row - min_row + 1,
        column_count=max_col - min_col + 1,
        row_numbers=row_numbers,
        column_numbers=column_numbers,
        rows=rows,
        dense_score=1.0,
        native_table_name=table.displayName,
        metadata={"table_ref": table.ref},
    )


def _score_rows_with_repeated_shape(region: SheetRegion) -> float:
    """Score how consistently populated each region row is."""
    widths = [len(row) for row in region.rows if row]
    if not widths:
        return 0.0
    ordered = sorted(widths)
    median_width = ordered[len(ordered) // 2]
    matching = sum(1 for width in widths if abs(width - median_width) <= 1)
    return matching / len(widths)


def _estimate_long_text_ratio(region: SheetRegion) -> float:
    """Estimate how much of the region reads like narrative text.

    Uses median cell length to avoid penalizing tables that have
    one or two description columns alongside short-value columns.
    """
    values = [value for row in region.rows for value in row.values()]
    if not values:
        return 1.0
    lengths = sorted(len(v) for v in values)
    median_length = lengths[len(lengths) // 2]
    if median_length >= 80:
        return 1.0
    if median_length <= 20:
        return 0.0
    return (median_length - 20) / 60.0


def score_dense_table_region(region: SheetRegion) -> float:
    """Heuristically score how much a region looks like a dense table.

    Params:
        region: Detected worksheet region

    Returns:
        float — score from 0.0 to 1.0

    Example:
        >>> region = SheetRegion(
        ...     region_id="r1", region_type="framing",
        ...     min_row=1, min_col=1, max_row=3, max_col=3,
        ...     non_empty_cells=9, row_count=3, column_count=3,
        ...     row_numbers=[1, 2, 3],
        ...     column_numbers=[1, 2, 3],
        ...     rows=[{1: "A", 2: "B", 3: "C"}] * 3,
        ... )
        >>> 0.0 <= score_dense_table_region(region) <= 1.0
        True
    """
    if region.region_type == "native_table":
        return 1.0
    if region.row_count < 2 or region.column_count < 2:
        return 0.0
    shape_score = _score_rows_with_repeated_shape(region)
    fill_ratio = region.non_empty_cells / (
        region.row_count * region.column_count
    )
    column_factor = min(1.0, region.column_count / 3.0)
    raw_ratio = min(region.row_count / max(1, region.column_count), 2.0) / 2.0
    row_bias = raw_ratio * column_factor
    header_width = len(region.rows[0]) if region.rows else 0
    header_score = min(1.0, header_width / max(1, region.column_count))
    long_text_penalty = _estimate_long_text_ratio(region)
    score = (
        0.35 * shape_score
        + 0.25 * fill_ratio
        + 0.2 * row_bias
        + 0.1 * header_score
        + 0.1 * (1.0 - long_text_penalty)
    )
    if region.row_count >= 3:
        score += 0.05
    return round(min(0.99, max(0.0, score)), 4)


def _label_regions(regions: list[SheetRegion]) -> list[SheetRegion]:
    """Assign IDs and dense-table labels to detected regions."""
    ordered = sorted(regions, key=lambda item: (item.min_row, item.min_col))
    labeled: list[SheetRegion] = []
    for index, region in enumerate(ordered, start=1):
        region.region_id = f"region_{index}"
        region.dense_score = score_dense_table_region(region)
        if region.region_type != "native_table":
            region.region_type = (
                "dense_table_candidate"
                if region.dense_score >= 0.55
                else "framing"
            )
        labeled.append(region)
    return labeled


def collect_sheet_cells(
    sheet: Worksheet,
    cached_values: dict[tuple[int, int], Any],
) -> list[SheetCell]:
    """Collect populated worksheet cells with preserved coordinates.

    Params:
        sheet: Source worksheet
        cached_values: Data-only workbook values keyed by row/column

    Returns:
        list[SheetCell] for every populated cell

    Example:
        >>> workbook = Workbook()
        >>> sheet = workbook.active
        >>> sheet["A1"] = "Title"
        >>> len(collect_sheet_cells(sheet, {}))
        1
    """
    cells: list[SheetCell] = []
    for row in sheet.iter_rows():
        for cell in row:
            raw_value = cached_values.get((cell.row, cell.column), cell.value)
            value = _normalize_cell_value(raw_value)
            if not value:
                continue
            cells.append(
                SheetCell(
                    row=cell.row,
                    column=cell.column,
                    value=value,
                    is_formula=(
                        isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    ),
                )
            )
    return cells


def detect_native_table_regions(
    sheet: Worksheet,
    cached_values: dict[tuple[int, int], Any],
) -> list[SheetRegion]:
    """Build regions from native Excel table definitions."""
    regions: list[SheetRegion] = []
    for table in _iter_native_tables(sheet):
        regions.append(_build_native_region(sheet, cached_values, table))
    return _label_regions(regions)


def detect_sheet_regions(cells: list[SheetCell]) -> list[SheetRegion]:
    """Detect populated regions using blank row and column bands.

    Params:
        cells: Populated worksheet cells

    Returns:
        list[SheetRegion] for non-native regions

    Example:
        >>> detect_sheet_regions([])
        []
    """
    if not cells:
        return []
    regions = [
        _build_region(group) for group in _split_into_leaf_groups(cells)
    ]
    return _label_regions(regions)


def build_sheet_regions(
    sheet: Worksheet,
    cached_values: dict[tuple[int, int], Any],
) -> list[SheetRegion]:
    """Build native and derived regions for a worksheet."""
    native_regions = detect_native_table_regions(sheet, cached_values)
    all_cells = collect_sheet_cells(sheet, cached_values)
    covered_positions = {
        (row_number, column_number)
        for region in native_regions
        for row_number in range(region.min_row, region.max_row + 1)
        for column_number in range(region.min_col, region.max_col + 1)
    }
    remaining_cells = [
        cell
        for cell in all_cells
        if (cell.row, cell.column) not in covered_positions
    ]
    combined = native_regions + detect_sheet_regions(remaining_cells)
    return _label_regions(combined)


def select_dense_table_region(
    regions: list[SheetRegion],
) -> SheetRegion | None:
    """Select the best dense-table candidate from detected regions.

    Params:
        regions: Candidate worksheet regions

    Returns:
        SheetRegion or None when no region scores high enough

    Example:
        >>> select_dense_table_region([])
        None
    """
    if not regions:
        return None

    def _selection_bucket(region: SheetRegion) -> int:
        """Bucket dense scores so size breaks near-ties."""
        return int(region.dense_score * 10 + 0.5)

    ordered = sorted(
        regions,
        key=lambda item: (
            _selection_bucket(item),
            item.non_empty_cells,
            item.row_count,
            item.column_count,
            item.region_type == "native_table",
            item.dense_score,
        ),
        reverse=True,
    )
    selected = ordered[0]
    if selected.dense_score < 0.55:
        return None
    return selected


def build_region_used_range(region: SheetRegion) -> str:
    """Build Excel A1 notation for a region's exact bounds.

    Params:
        region: Worksheet region

    Returns:
        str — A1 range

    Example:
        >>> build_region_used_range(
        ...     SheetRegion(
        ...         region_id="r1", region_type="framing",
        ...         min_row=1, min_col=1, max_row=2, max_col=2,
        ...         non_empty_cells=4, row_count=2, column_count=2,
        ...         row_numbers=[1, 2], column_numbers=[1, 2],
        ...         rows=[{1: "A", 2: "B"}, {1: "C", 2: "D"}],
        ...     )
        ... )
        'A1:B2'
    """
    start = f"{get_column_letter(region.min_col)}{region.min_row}"
    end = f"{get_column_letter(region.max_col)}{region.max_row}"
    if start == end:
        return start
    return f"{start}:{end}"
